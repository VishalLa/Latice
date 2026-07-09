
from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import Any, Optional

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from schema import LedgerFormat, BankStatement

C_HEADER_BG   = "1E3A5F"
C_HEADER_FG   = "FFFFFF"
C_SUBHEADER   = "2E6DA4"
C_ROW_ALT     = "EBF3FB"
C_TOTAL_BG    = "FFF2CC"
C_SECTION_BG  = "D9E1F2"
C_BORDER      = "BFBFBF"
C_GREEN_BG    = "E2EFDA"
C_AMBER_BG    = "FFF2CC"
C_RED_BG      = "FFCCCC"
C_ORANGE_BG   = "FCE4D6"
C_PURPLE_BG   = "EAD1DC"
C_TEAL_BG     = "D0E4E4"
C_BLUE_BG     = "DAEEF3"
C_GREY_BG     = "F2F2F2"

INR = "\u20b9#,##0.00"

_CONF_COLOURS = {
    "High":   C_GREEN_BG,
    "Medium": C_AMBER_BG,
    "Low":    C_ORANGE_BG,
}

_CONF_FLOAT_BG = [
    (0.80, C_GREEN_BG),
    (0.50, C_AMBER_BG),
    (0.00, C_ORANGE_BG),
]

def _conf_bg(conf: Any) -> Optional[str]:
    if isinstance(conf, (float, int)) and not isinstance(conf, bool):
        for threshold, colour in _CONF_FLOAT_BG:
            if conf >= threshold:
                return colour
        return C_ORANGE_BG
    return _CONF_COLOURS.get(str(conf), None)

def _border(style="thin", color=C_BORDER):
    s = Side(style=style, color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def _cell(ws, row, col, value=None, *, bold=False, size=9, bg=None,
          fg="000000", align="left", num_fmt=None, italic=False,
          border=None, wrap=False):
    c = ws.cell(row=row, column=col, value=value)
    c.font      = Font(name="Arial", bold=bold, size=size, color=fg, italic=italic)
    c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
    if bg:
        c.fill = PatternFill("solid", fgColor=bg)
    if num_fmt:
        c.number_format = num_fmt
    if border:
        c.border = border
    return c

def _title(ws, row: int, text: str, n_cols: int, subtitle: str = "") -> int:
    ws.row_dimensions[row].height = 32
    ws.merge_cells(f"A{row}:{get_column_letter(n_cols)}{row}")
    _cell(ws, row, 1, text, bold=True, size=14,
          bg=C_HEADER_BG, fg=C_HEADER_FG, align="center")
    row += 1
    if subtitle:
        ws.row_dimensions[row].height = 18
        ws.merge_cells(f"A{row}:{get_column_letter(n_cols)}{row}")
        _cell(ws, row, 1, subtitle, size=9,
              bg=C_SUBHEADER, fg=C_HEADER_FG, align="center", italic=True)
        row += 1
    return row + 1

def _col_headers(ws, row: int, headers: list,
                 bg=C_SUBHEADER, fg=C_HEADER_FG) -> int:
    ws.row_dimensions[row].height = 22
    for c, h in enumerate(headers, 1):
        _cell(ws, row, c, h, bold=True, size=9,
              bg=bg, fg=fg, align="center", border=_border())
    return row + 1

def _empty_notice(ws, row: int, text: str, n_cols: int) -> int:
    ws.merge_cells(f"A{row}:{get_column_letter(n_cols)}{row}")
    _cell(ws, row, 1, text, size=9, italic=True, fg="888888", align="center")
    return row + 1

def _set_col_widths(ws, widths: list) -> None:
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def _ledger_row(rec: LedgerFormat) -> tuple:
    amount    = rec.debit_amount or rec.credit_amount
    direction = "Debit" if rec.debit_amount > 0 else "Credit"
    return (
        rec.ledger_id,
        rec.transaction_date or rec.transaction_date_raw or "",
        rec.account_name or "",
        rec.vendor_name  or "",
        rec.voucher_type or "",
        rec.reference_id or "",
        amount,
        direction,
        rec.source.value if getattr(rec, "source", None) else "",
    )

def _bank_row(rec: BankStatement) -> tuple:
    amount    = rec.debit or rec.credit
    direction = "Debit" if rec.debit > 0 else "Credit"
    return (
        str(rec.row_index),
        rec.date or rec.date_raw or "",
        rec.narration or "",
        rec.txn_id or "",
        amount,
        direction,
        f"\u20b9{rec.balance:,.2f}" if rec.balance is not None else "",
    )

def _ledger_ids_str(components: Any) -> str:
    if components is None:
        return ""
    if isinstance(components, str):
        return components
    if isinstance(components, list):
        ids = []
        for c in components:
            if isinstance(c, dict):
                ids.append(str(c.get("ledger_id", "")))
            else:
                ids.append(str(c))
        return " & ".join(ids)
    return str(components)

def _write_summary(wb, result: dict) -> None:
    ws = wb.create_sheet("Summary")
    ws.sheet_view.showGridLines = False
    _set_col_widths(ws, [34, 14, 30])

    stamp     = datetime.now().strftime("%d %B %Y, %I:%M %p")
    bank_name = result.get("bank_name") or "-"
    r = _title(ws, 1, "BANK RECONCILIATION STATEMENT", 3,
               f"Bank: {bank_name}  |  Generated: {stamp}")

    s   = result.get("summary", {}) or {}
    mq  = s.get("match_quality") or {}
    mem = s.get("memory") or {}
    rr  = s.get("residual_reconciliation") or {}

    rows = [
        ("RECORDS LOADED",                       None, None),
        ("Ledger records",                       s.get("ledger_records", 0),      C_SECTION_BG),
        ("Bank statement rows",                  s.get("bank_records",   0),      C_SECTION_BG),
        (None, None, None),

        ("MATCHING - ALL PHASES",                None, None),
        ("Phase 1 - Exact matches",               s.get("exact_matches",   0),    C_GREEN_BG),
        ("Phase 2 - Fuzzy matches",               s.get("fuzzy_matches",   0),    C_GREEN_BG),
        ("Recurring-pattern (memory) matches",     s.get("memory_matches",  0),    C_GREEN_BG),
        ("Phase 3 - AI semantic matches",          s.get("ai_matches",      0),    C_GREEN_BG),
        ("Phase 3 - AI audit queue (low conf.)",   s.get("ai_audit_queue",  0),    C_AMBER_BG),
        (None, None, None),

        ("PHASE 4 - RESIDUAL RECONCILIATION",     None, None),
        ("Timing re-matches resolved",            rr.get("timing_resolved",  0),  C_TEAL_BG),
        ("Split (subset-sum) matches resolved",    rr.get("splits_resolved",  0),  C_TEAL_BG),
        ("Journal drafts generated",               rr.get("drafts_generated", 0),  C_AMBER_BG),
        ("Still requiring human review",           rr.get("still_unresolved", 0),  C_ORANGE_BG),
        ("Journal drafting used AI",
         "No - heuristic fallback" if rr.get("ai_skipped") else "Yes",             C_GREY_BG),
        (None, None, None),

        ("MATCH QUALITY",                         None, None),
        ("Average confidence (all matches)",
         f"{mq.get('average_confidence'):.1%}" if mq.get("average_confidence") is not None else "-",
         C_BLUE_BG),
        ("Matches below quality bar",
         f"{mq.get('low_confidence_match_count', 0)} ({mq.get('low_confidence_match_pct', 0.0)}%)",
         C_AMBER_BG),
        (None, None, None),

        ("RECURRING-PATTERN MEMORY",              None, None),
        ("Patterns recognised this run",
         mem.get("patterns_matched", mem.get("recognised", "-")) if mem else "No memory store supplied",
         C_PURPLE_BG),
        (None, None, None),

        ("FINAL UNRECONCILED (after all phases)", None, None),
        ("Unreconciled ledger rows",               s.get("unreconciled_ledger", 0), C_RED_BG),
        ("Unreconciled bank rows",                 s.get("unreconciled_bank",   0), C_RED_BG),
        (None, None, None),

        ("AI phase (Phase 3) skipped",
         "Yes" if s.get("ai_skipped") else "No",                                  C_GREY_BG),
        ("Warnings raised",                        len(result.get("warnings", [])), C_AMBER_BG),
    ]

    for label, value, bg in rows:
        ws.row_dimensions[r].height = 20
        if label is None:
            r += 1
            continue
        if value is None:
            ws.merge_cells(f"A{r}:C{r}")
            _cell(ws, r, 1, label, bold=True, size=10,
                  bg=C_HEADER_BG, fg=C_HEADER_FG, border=_border())
        else:
            _cell(ws, r, 1, f"  {label}", size=10, bg=bg, border=_border())
            _cell(ws, r, 2, value, align="center", bold=True, size=11, bg=bg, border=_border())
            _cell(ws, r, 3, "", bg=bg, border=_border())
        r += 1

def _write_exact_matches(wb, matches: list) -> None:
    ws = wb.create_sheet("Exact Matches")
    ws.sheet_view.showGridLines = False
    _set_col_widths(ws, [4, 12, 12, 14, 14, 18, 20])

    r = _title(ws, 1, "PHASE 1 - EXACT MATCHES", 7,
               f"{len(matches)} exact matches (date + amount + optional reference/UTR)")
    r = _col_headers(ws, r, [
        "#", "Ledger ID", "Bank Row", "Date", "Amount (₹)",
        "Reference Matched", "Confirmed Via",
    ])
    ws.freeze_panes = f"A{r}"

    if not matches:
        return _empty_notice(ws, r, "No exact matches found.", 7)

    for i, m in enumerate(matches, 1):
        alt = C_ROW_ALT if i % 2 == 0 else None
        bg  = C_GREEN_BG if m.get("reference_matched") else alt
        ws.row_dimensions[r].height = 18
        _cell(ws, r, 1, i,                                 align="center", size=8, bg=bg, border=_border())
        _cell(ws, r, 2, str(m.get("ledger_id", "")),       bg=bg, border=_border())
        _cell(ws, r, 3, str(m.get("bank_id",   "")),       bg=bg, border=_border())
        _cell(ws, r, 4, str(m.get("date",      "")),       align="center", bg=bg, border=_border())
        _cell(ws, r, 5, m.get("amount"),                   align="right", num_fmt=INR, bg=bg, border=_border())
        _cell(ws, r, 6, "Yes ✓" if m.get("reference_matched") else "No - amount+date only",
              align="center", bg=bg, border=_border())
        _cell(ws, r, 7, m.get("confirmation_method", ""),  align="center", size=8, bg=bg, border=_border())
        r += 1

def _write_fuzzy_matches(wb, matches: list) -> None:
    ws = wb.create_sheet("Fuzzy Matches")
    ws.sheet_view.showGridLines = False
    _set_col_widths(ws, [4, 14, 14, 30, 12, 50])

    r = _title(ws, 1, "PHASE 2 - FUZZY MATCHES", 6,
               f"{len(matches)} matches across the fuzzy-matching strategy set")
    r = _col_headers(ws, r, [
        "#", "Ledger ID", "Bank Row", "Adjustment Type", "Confidence", "Details",
    ])
    ws.freeze_panes = f"A{r}"

    if not matches:
        return _empty_notice(ws, r, "No fuzzy matches found.", 6)

    for i, m in enumerate(matches, 1):
        conf = m.get("confidence_score", "")
        bg   = _conf_bg(conf) or (C_ROW_ALT if i % 2 == 0 else None)
        ws.row_dimensions[r].height = 18
        _cell(ws, r, 1, i,                                        align="center", size=8, bg=bg, border=_border())
        _cell(ws, r, 2, str(m.get("ledger_id", "") or "-"),       bg=bg, border=_border())
        _cell(ws, r, 3, str(m.get("bank_id",   "") or "-"),       bg=bg, border=_border())
        _cell(ws, r, 4, m.get("adjustment_type", ""),             bg=bg, border=_border())
        _cell(ws, r, 5, str(conf), align="center", bold=True,     bg=bg, border=_border())
        _cell(ws, r, 6, m.get("details", ""), size=8, italic=True, wrap=True, bg=bg, border=_border())
        r += 1

def _write_memory_matches(wb, matches: list) -> None:
    ws = wb.create_sheet("Memory Matches")
    ws.sheet_view.showGridLines = False
    _set_col_widths(ws, [4, 14, 14, 34, 12, 50])

    r = _title(ws, 1, "RECURRING-PATTERN (MEMORY) MATCHES", 6,
               f"{len(matches)} matches recognised from prior runs, before the AI phase ran")
    r = _col_headers(ws, r, [
        "#", "Ledger ID", "Bank Row", "Pattern", "Confidence", "Details",
    ])
    ws.freeze_panes = f"A{r}"

    if not matches:
        return _empty_notice(ws, r, "No memory matches this run - either no memory store was "
                                     "supplied, or no recurring counterparty was recognised.", 6)

    for i, m in enumerate(matches, 1):
        conf = m.get("confidence_score", "")
        bg   = _conf_bg(conf) or (C_ROW_ALT if i % 2 == 0 else None)
        ws.row_dimensions[r].height = 18
        _cell(ws, r, 1, i,                                    align="center", size=8, bg=bg, border=_border())
        _cell(ws, r, 2, str(m.get("ledger_id", "") or "-"),   bg=bg, border=_border())
        _cell(ws, r, 3, str(m.get("bank_id",   "") or "-"),   bg=bg, border=_border())
        _cell(ws, r, 4, m.get("adjustment_type", ""),         bg=bg, border=_border())
        _cell(ws, r, 5, str(conf), align="center", bold=True, bg=bg, border=_border())
        _cell(ws, r, 6, m.get("details", ""), size=8, italic=True, wrap=True, bg=bg, border=_border())
        r += 1

def _write_ai_matches(wb, matches: list) -> None:
    ws = wb.create_sheet("AI Matches")
    ws.sheet_view.showGridLines = False
    _set_col_widths(ws, [4, 14, 16, 12, 50])

    r = _title(ws, 1, "PHASE 3 - AI SEMANTIC MATCHES", 5,
               f"{len(matches)} semantic matches identified by the LLM (1:1 and many:1)")
    r = _col_headers(ws, r, [
        "#", "Ledger ID(s)", "Bank Row", "Confidence", "Reasoning",
    ])
    ws.freeze_panes = f"A{r}"

    if not matches:
        return _empty_notice(ws, r, "No AI matches found (or the AI phase was skipped - "
                                     "see Summary for the reason).", 5)

    for i, m in enumerate(matches, 1):
        conf = m.get("confidence", "")
        bg   = _conf_bg(conf) or (C_ROW_ALT if i % 2 == 0 else None)
        # AI1to1Match has ledger_id (str); AIManyToOneMatch has ledger_ids (list)
        ledger_display = (
            m.get("ledger_id") if "ledger_id" in m
            else _ledger_ids_str(m.get("ledger_ids"))
        )
        ws.row_dimensions[r].height = 18
        _cell(ws, r, 1, i,                                    align="center", size=8, bg=bg, border=_border())
        _cell(ws, r, 2, str(ledger_display or "-"),           bg=bg, border=_border())
        _cell(ws, r, 3, str(m.get("bank_id", "") or "-"),     bg=bg, border=_border())
        _cell(ws, r, 4, f"{conf:.0%}" if isinstance(conf, (int, float)) else str(conf),
              align="center", bold=True, bg=bg, border=_border())
        _cell(ws, r, 5, m.get("reasoning", ""), size=8, italic=True, wrap=True, bg=bg, border=_border())
        r += 1

def _write_ai_audit_queue(wb, queue: list) -> None:
    ws = wb.create_sheet("AI Audit Queue")
    ws.sheet_view.showGridLines = False
    _set_col_widths(ws, [4, 14, 16, 12, 22, 50])

    r = _title(ws, 1, "PHASE 3 - AI AUDIT QUEUE", 6,
               f"{len(queue)} AI-proposed matches below the confidence threshold - human review required")
    r = _col_headers(ws, r, [
        "#", "Ledger ID(s)", "Bank Row", "Confidence", "Action Required", "Reasoning",
    ])
    ws.freeze_panes = f"A{r}"

    if not queue:
        return _empty_notice(ws, r, "No items in the AI audit queue.", 6)

    for i, item in enumerate(queue, 1):
        conf = item.get("confidence", "")
        bg   = _conf_bg(conf) or C_AMBER_BG
        ledger_display = (
            item.get("ledger_id") if "ledger_id" in item
            else _ledger_ids_str(item.get("ledger_ids"))
        )
        ws.row_dimensions[r].height = 20
        _cell(ws, r, 1, i,                                align="center", size=8, bg=bg, border=_border())
        _cell(ws, r, 2, str(ledger_display or "-"),       bg=bg, border=_border())
        _cell(ws, r, 3, str(item.get("bank_id", "") or "-"), bg=bg, border=_border())
        _cell(ws, r, 4, f"{conf:.0%}" if isinstance(conf, (int, float)) else str(conf),
              align="center", bold=True, bg=bg, border=_border())
        _cell(ws, r, 5, item.get("action", ""), size=8, bg=bg, border=_border())
        _cell(ws, r, 6, item.get("reasoning", ""), size=8, italic=True, wrap=True, bg=bg, border=_border())
        r += 1

def _write_timing_matches(wb, matches: list) -> None:
    ws = wb.create_sheet("Residual Timing Matches")
    ws.sheet_view.showGridLines = False
    _set_col_widths(ws, [4, 14, 12, 14, 12, 12, 50])

    r = _title(ws, 1, "PHASE 4 · STEP 2 - TIMING RE-MATCHES", 7,
               f"{len(matches)} matches found via a wide-window (45-day) exact-amount re-check")
    r = _col_headers(ws, r, [
        "#", "Ledger ID", "Bank Row", "Amount (₹)", "Date Gap (days)", "Confidence", "Details",
    ])
    ws.freeze_panes = f"A{r}"

    if not matches:
        return _empty_notice(ws, r, "No timing candidates resolved this run.", 7)

    for i, m in enumerate(matches, 1):
        conf       = m.get("confidence_score", "")
        ambiguous  = m.get("ambiguous", False)
        bg         = C_AMBER_BG if ambiguous else (_conf_bg(conf) or (C_ROW_ALT if i % 2 == 0 else None))
        ws.row_dimensions[r].height = 20
        _cell(ws, r, 1, i,                                     align="center", size=8, bg=bg, border=_border())
        _cell(ws, r, 2, str(m.get("ledger_id", "")),           bg=bg, border=_border())
        _cell(ws, r, 3, str(m.get("bank_id",   "")),           bg=bg, border=_border())
        _cell(ws, r, 4, m.get("amount"),                       align="right", num_fmt=INR, bg=bg, border=_border())
        _cell(ws, r, 5, m.get("date_gap_days"),                align="center", bg=bg, border=_border())
        _cell(ws, r, 6, ("⚠ " if ambiguous else "") + str(conf),
              align="center", bold=True, bg=bg, border=_border())
        _cell(ws, r, 7, m.get("details", ""), size=8, italic=True, wrap=True, bg=bg, border=_border())
        r += 1

def _write_split_matches(wb, matches: list) -> None:
    ws = wb.create_sheet("Residual Split Matches")
    ws.sheet_view.showGridLines = False
    _set_col_widths(ws, [4, 24, 14, 14, 12, 50])

    r = _title(ws, 1, "PHASE 4 · STEP 3 - SPLIT (SUBSET-SUM) MATCHES", 6,
               f"{len(matches)} matches where one side's amount equals a sum of entries on the other")
    r = _col_headers(ws, r, [
        "#", "Ledger ID(s)", "Bank Row(s)", "Amount (₹)", "Confidence", "Details",
    ])
    ws.freeze_panes = f"A{r}"

    if not matches:
        return _empty_notice(ws, r, "No split candidates resolved this run.", 6)

    for i, m in enumerate(matches, 1):
        conf      = m.get("confidence_score", "")
        ambiguous = m.get("ambiguous", False)
        bg        = C_AMBER_BG if ambiguous else (_conf_bg(conf) or (C_ROW_ALT if i % 2 == 0 else None))

        ledger_display = (
            _ledger_ids_str(m["ledger_components"]) if "ledger_components" in m
            else str(m.get("ledger_id", ""))
        )
        bank_display = (
            str(m.get("bank_id", "")) if "bank_components" not in m
            else " & ".join(str(c.get("bank_id", "")) for c in m["bank_components"])
        )

        ws.row_dimensions[r].height = 20
        _cell(ws, r, 1, i,                 align="center", size=8, bg=bg, border=_border())
        _cell(ws, r, 2, ledger_display,     wrap=True, bg=bg, border=_border())
        _cell(ws, r, 3, bank_display,       wrap=True, bg=bg, border=_border())
        _cell(ws, r, 4, m.get("amount"),   align="right", num_fmt=INR, bg=bg, border=_border())
        _cell(ws, r, 5, ("⚠ " if ambiguous else "") + str(conf),
              align="center", bold=True, bg=bg, border=_border())
        _cell(ws, r, 6, m.get("details", ""), size=8, italic=True, wrap=True, bg=bg, border=_border())
        r += 1

def _write_suggested_journal_entries(wb, entries: list) -> None:
    ws = wb.create_sheet("Suggested Journal Entries")
    ws.sheet_view.showGridLines = False
    _set_col_widths(ws, [4, 12, 12, 14, 22, 22, 34, 12, 12, 14])

    r = _title(ws, 1, "PHASE 4 · STEP 4 - SUGGESTED JOURNAL ENTRIES", 10,
               f"{len(entries)} draft entries for accountant review. Set the Status column to "
               f"APPROVED or MODIFIED, then post via journal_posting.approve_journal_entries().")
    r = _col_headers(ws, r, [
        "#", "Bank Row", "Date", "Amount (₹)", "Debit A/c", "Credit A/c",
        "Narration", "Confidence", "Source", "Status",
    ])
    ws.freeze_panes = f"A{r}"

    if not entries:
        return _empty_notice(ws, r, "No missing-entry candidates this run - nothing to draft.", 10)

    for i, e in enumerate(entries, 1):
        conf = e.get("confidence", 0.0)
        bg   = _conf_bg(conf) or (C_ROW_ALT if i % 2 == 0 else None)
        ws.row_dimensions[r].height = 20
        _cell(ws, r, 1,  i,                                 align="center", size=8, bg=bg, border=_border())
        _cell(ws, r, 2,  str(e.get("bank_id", "")),         align="center", bg=bg, border=_border())
        _cell(ws, r, 3,  str(e.get("date", "")),            align="center", bg=bg, border=_border())
        _cell(ws, r, 4,  e.get("amount"),                   align="right", num_fmt=INR, bg=bg, border=_border())
        _cell(ws, r, 5,  e.get("debit_account", ""),        bg=bg, border=_border())
        _cell(ws, r, 6,  e.get("credit_account", ""),       bg=bg, border=_border())
        _cell(ws, r, 7,  e.get("entry_narrative", ""),      size=8, wrap=True, bg=bg, border=_border())
        _cell(ws, r, 8,  f"{conf:.0%}" if isinstance(conf, (int, float)) else str(conf),
              align="center", bold=True, bg=bg, border=_border())
        _cell(ws, r, 9,  e.get("source", ""),                align="center", size=8, italic=True, bg=bg, border=_border())
        _cell(ws, r, 10, e.get("status", "pending_review"),  align="center", bold=True, bg=C_AMBER_BG, border=_border())
        r += 1

def _write_human_review_queue(wb, queue: list) -> None:
    ws = wb.create_sheet("Human Review Queue")
    ws.sheet_view.showGridLines = False
    _set_col_widths(ws, [4, 8, 12, 14, 30, 12, 40, 34, 12])

    r = _title(ws, 1, "PHASE 4 · STEP 5 - HUMAN REVIEW QUEUE", 9,
               f"{len(queue)} items with no confident automated resolution - reviewed with the "
               f"3 closest candidates from the other side attached")
    r = _col_headers(ws, r, [
        "#", "Side", "Record ID", "Amount (₹)", "Narration / Account",
        "Date", "Classification", "Closest Candidates", "Suggested Action",
    ])
    ws.freeze_panes = f"A{r}"

    if not queue:
        return _empty_notice(ws, r, "Nothing pending human review this run.", 9)

    for i, item in enumerate(queue, 1):
        bg = C_RED_BG if item.get("side") == "bank" else C_ORANGE_BG
        candidates = item.get("closest_candidates") or []
        cand_text = "; ".join(
            f"{c.get('candidate_id')} (Δ₹{c.get('amount_difference')}, "
            f"{c.get('date_gap_days')}d, sim {c.get('text_similarity')})"
            for c in candidates
        ) or "-"

        ws.row_dimensions[r].height = 34
        _cell(ws, r, 1, i,                                   align="center", size=8, bg=bg, border=_border())
        _cell(ws, r, 2, item.get("side", "").capitalize(),   align="center", bold=True, bg=bg, border=_border())
        _cell(ws, r, 3, str(item.get("id", "")),             align="center", bg=bg, border=_border())
        _cell(ws, r, 4, item.get("amount"),                  align="right", num_fmt=INR, bg=bg, border=_border())
        _cell(ws, r, 5, item.get("narration_or_account", ""), size=8, wrap=True, bg=bg, border=_border())
        _cell(ws, r, 6, str(item.get("date", "")),           align="center", bg=bg, border=_border())
        _cell(ws, r, 7, item.get("classification", ""),      align="center", size=8, bg=bg, border=_border())
        _cell(ws, r, 8, cand_text,                            size=7, italic=True, wrap=True, bg=bg, border=_border())
        _cell(ws, r, 9, item.get("suggested_action", ""),    size=8, wrap=True, bg=bg, border=_border())
        r += 1

def _write_audit_investigation(wb, items: list) -> None:
    ws = wb.create_sheet("Audit Investigation")
    ws.sheet_view.showGridLines = False
    _set_col_widths(ws, [4, 12, 34, 14, 10, 40, 40])

    r = _title(ws, 1, "AUDIT INVESTIGATION - FLAGGED BANK REVERSALS", 7,
               f"{len(items)} bank rows flagged as likely reversals; not force-matched, "
               f"require a manual General Ledger journal entry")
    r = _col_headers(ws, r, [
        "#", "Bank Row", "Narration", "Amount (₹)", "Direction", "Flag Reason", "Action Required",
    ])
    ws.freeze_panes = f"A{r}"

    if not items:
        return _empty_notice(ws, r, "No audit-investigation items this run.", 7)

    for i, item in enumerate(items, 1):
        bg = C_ORANGE_BG if i % 2 == 0 else C_RED_BG
        ws.row_dimensions[r].height = 20
        _cell(ws, r, 1, i,                                          align="center", size=8, bg=bg, border=_border())
        _cell(ws, r, 2, str(getattr(item, "bank_row_index", "")),   align="center", bg=bg, border=_border())
        _cell(ws, r, 3, getattr(item, "narration", ""),             size=8, wrap=True, bg=bg, border=_border())
        _cell(ws, r, 4, getattr(item, "amount", None),              align="right", num_fmt=INR, bg=bg, border=_border())
        _cell(ws, r, 5, getattr(item, "direction", "").capitalize(),align="center", bg=bg, border=_border())
        _cell(ws, r, 6, getattr(item, "flag_reason", ""),           size=8, wrap=True, bg=bg, border=_border())
        _cell(ws, r, 7, getattr(item, "action_required", ""),       size=8, italic=True, wrap=True, bg=bg, border=_border())
        r += 1

def _write_ignored_metadata(wb, items: list) -> None:
    ws = wb.create_sheet("Ignored Metadata")
    ws.sheet_view.showGridLines = False
    _set_col_widths(ws, [4, 10, 12, 40, 50])

    r = _title(ws, 1, "IGNORED METADATA / ZERO-AMOUNT ROWS", 5,
               f"{len(items)} rows silently dropped before matching began "
               f"(header rows, zero-amount metadata - not real transactions)")
    r = _col_headers(ws, r, ["#", "Source", "Row Ref", "Narration", "Reason"])
    ws.freeze_panes = f"A{r}"

    if not items:
        return _empty_notice(ws, r, "No rows were ignored this run.", 5)

    for i, item in enumerate(items, 1):
        bg = C_GREY_BG if i % 2 == 0 else None
        ws.row_dimensions[r].height = 18
        _cell(ws, r, 1, i,                                       align="center", size=8, bg=bg, border=_border())
        _cell(ws, r, 2, getattr(item, "source", "").capitalize(),align="center", bg=bg, border=_border())
        _cell(ws, r, 3, str(getattr(item, "row_ref", "")),       align="center", bg=bg, border=_border())
        _cell(ws, r, 4, getattr(item, "narration", ""),          size=8, bg=bg, border=_border())
        _cell(ws, r, 5, getattr(item, "reason", ""),             size=8, italic=True, wrap=True, bg=bg, border=_border())
        r += 1

def _write_unreconciled(wb, unreconciled: dict) -> None:
    ws = wb.create_sheet("Unreconciled")
    ws.sheet_view.showGridLines = False

    ledger_items = unreconciled.get("ledger", []) or []
    bank_items   = unreconciled.get("bank",   []) or []

    _set_col_widths(ws, [12, 12, 26, 20, 16, 16, 14, 10, 10])
    r = _title(ws, 1, "FINAL UNRECONCILED ITEMS", 9,
               f"{len(ledger_items)} ledger row(s), {len(bank_items)} bank row(s) - "
               f"nothing left could be resolved automatically or drafted")

    r = _section_header_local(ws, r, "UNRECONCILED LEDGER ENTRIES", 9)
    r = _col_headers(ws, r, [
        "Ledger ID", "Date", "Account Name", "Vendor", "Voucher Type",
        "Reference", "Amount (₹)", "Dir.", "Source",
    ])
    if not ledger_items:
        r = _empty_notice(ws, r, "No unreconciled ledger entries - fully matched.", 9)
    else:
        for i, rec in enumerate(ledger_items):
            bg = C_ROW_ALT if i % 2 == 0 else None
            vals = _ledger_row(rec)
            ws.row_dimensions[r].height = 18
            for c, v in enumerate(vals, 1):
                align = "right" if c == 7 else ("center" if c in (1, 2, 8, 9) else "left")
                num_fmt = INR if c == 7 else None
                _cell(ws, r, c, v, align=align, num_fmt=num_fmt, size=8, bg=bg, border=_border())
            r += 1

    r += 1
    r = _section_header_local(ws, r, "UNRECONCILED BANK ROWS", 9)
    r = _col_headers(ws, r, [
        "Row #", "Date", "Narration", "Txn/UTR ID", "", "Amount (₹)", "Dir.", "Balance", "",
    ])
    if not bank_items:
        r = _empty_notice(ws, r, "No unreconciled bank rows - fully matched.", 9)
    else:
        for i, rec in enumerate(bank_items):
            bg = C_ROW_ALT if i % 2 == 0 else None
            row = _bank_row(rec)
            ws.row_dimensions[r].height = 18
            mapped = [row[0], row[1], row[2], row[3], "", row[4], row[5], row[6], ""]
            for c, v in enumerate(mapped, 1):
                align = "right" if c == 6 else ("center" if c in (1, 2, 7) else "left")
                num_fmt = INR if c == 6 else None
                _cell(ws, r, c, v, align=align, num_fmt=num_fmt, size=8, bg=bg, border=_border())
            r += 1

def _section_header_local(ws, row: int, text: str, n_cols: int) -> int:
    ws.row_dimensions[row].height = 20
    ws.merge_cells(f"A{row}:{get_column_letter(n_cols)}{row}")
    _cell(ws, row, 1, f"  {text}", bold=True, size=10, bg=C_HEADER_BG, fg=C_HEADER_FG, border=_border())
    return row + 1

def _write_warnings(wb, warnings: list) -> None:
    ws = wb.create_sheet("Warnings")
    ws.sheet_view.showGridLines = False
    _set_col_widths(ws, [6, 100])

    r = _title(ws, 1, "PIPELINE WARNINGS", 2, f"{len(warnings)} warning(s) raised during this run")
    r = _col_headers(ws, r, ["#", "Warning"])

    if not warnings:
        return _empty_notice(ws, r, "No warnings raised.", 2)

    for i, w in enumerate(warnings, 1):
        bg = C_AMBER_BG if i % 2 == 0 else C_ROW_ALT
        ws.row_dimensions[r].height = 30
        _cell(ws, r, 1, i, align="center", size=8, bg=bg, border=_border())
        _cell(ws, r, 2, str(w), size=9, wrap=True, bg=bg, border=_border())
        r += 1

def write_bank_recon_xlsx(
    recon_result: dict,
    gl_records:   list,
    bank_records: list,
    output_path:  "Path | str",
) -> None:
    path = Path(output_path)
    wb   = openpyxl.Workbook()
    wb.remove(wb.active)

    _write_summary(wb, recon_result)
    _write_exact_matches(wb,          recon_result.get("EXACT_MATCHES",  []))
    _write_fuzzy_matches(wb,          recon_result.get("FUZZY_MATCHES",  []))
    _write_memory_matches(wb,         recon_result.get("MEMORY_MATCHES", []))
    _write_ai_matches(wb,             recon_result.get("AI_MATCHES",     []))
    _write_ai_audit_queue(wb,         recon_result.get("AI_AUDIT_QUEUE", []))
    _write_timing_matches(wb,         recon_result.get("RESIDUAL_TIMING_MATCHES",   []))
    _write_split_matches(wb,          recon_result.get("RESIDUAL_SPLIT_MATCHES",    []))
    _write_suggested_journal_entries(wb, recon_result.get("SUGGESTED_JOURNAL_ENTRIES", []))
    _write_human_review_queue(wb,     recon_result.get("HUMAN_REVIEW_QUEUE", []))
    _write_audit_investigation(wb,    recon_result.get("AUDIT_INVESTIGATION", []))
    _write_ignored_metadata(wb,       recon_result.get("IGNORED_METADATA", []))
    _write_unreconciled(wb,           recon_result.get("UNRECONCILED_ITEMS", {"ledger": [], "bank": []}))
    _write_warnings(wb,               recon_result.get("warnings", []))

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(path))
