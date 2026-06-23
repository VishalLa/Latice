"""
GSTR-1 Outward Supplies Return — XLSX Export

Converts the dict produced by gstr1_builder.build_gstr1() into a
formatted Excel workbook.

Sheets produced:
  1. Cover          — filing period, totals at a glance
  2. B2B (Table 4)  — invoice-wise B2B taxable outward supplies
  3. B2C (Table 7)  — B2C large inter-state supplies (consolidated)
  4. Nil & Exempt (Table 8) — nil-rated / exempt / non-GST summary
  5. HSN Summary (Table 12) — HSN/SAC-wise outward supply summary

Usage
-----
    from gstr1_builder import build_gstr1
    from gstr1_xlsx    import write_gstr1_xlsx

    gstr1 = build_gstr1(bills, period_label="Apr-2025")
    write_gstr1_xlsx(gstr1, "GSTR1_Apr2025.xlsx")
"""

from __future__ import annotations

from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ── Colour palette (matches ledger/bank_recon xlsx palette) ──────────────────
C_HEADER_BG  = "1E3A5F"
C_HEADER_FG  = "FFFFFF"
C_SUBHEADER  = "2E6DA4"
C_ROW_ALT    = "EBF3FB"
C_TOTAL_BG   = "FFF2CC"
C_GRAND_BG   = "1E3A5F"
C_GRAND_FG   = "FFFFFF"
C_SECTION_BG = "D9E1F2"
C_BORDER     = "BFBFBF"
C_GREEN_BG   = "E2EFDA"
C_AMBER_BG   = "FFF2CC"
C_RED_BG     = "FFCCCC"
C_WARN_BG    = "FCE4D6"

INR = "\u20b9#,##0.00"


# ── Helpers ───────────────────────────────────────────────────────────────────

def _side(style="thin", color=C_BORDER):
    return Side(style=style, color=color)


def _border(style="thin"):
    s = _side(style)
    return Border(left=s, right=s, top=s, bottom=s)


def _thick_bottom():
    t = _side("thin")
    m = _side("medium", "000000")
    return Border(left=t, right=t, top=t, bottom=m)


def _cell(ws, row, col, value=None, *, bold=False, size=9, bg=None,
          fg="000000", align="left", num_fmt=None, italic=False,
          border=None, wrap=False):
    c = ws.cell(row=row, column=col, value=value)
    c.font      = Font(name="Arial", bold=bold, size=size, color=fg, italic=italic)
    c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
    if bg:
        c.fill = PatternFill("solid", fgColor=bg)
    if num_fmt:
        c.number_format = num_fmt
    if border:
        c.border = border
    return c


def _set_widths(ws, widths: list[float]) -> None:
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _title_block(ws, row: int, title: str, n_cols: int,
                 subtitle: str = "", table_ref: str = "") -> int:
    ws.row_dimensions[row].height = 30
    ws.merge_cells(f"A{row}:{get_column_letter(n_cols)}{row}")
    label = f"{table_ref}  {title}" if table_ref else title
    _cell(ws, row, 1, label.strip(), bold=True, size=13,
          bg=C_HEADER_BG, fg=C_HEADER_FG, align="center")
    row += 1
    if subtitle:
        ws.row_dimensions[row].height = 17
        ws.merge_cells(f"A{row}:{get_column_letter(n_cols)}{row}")
        _cell(ws, row, 1, subtitle, size=9, bg=C_SUBHEADER,
              fg=C_HEADER_FG, align="center", italic=True)
        row += 1
    return row + 1  # blank gap


def _col_headers(ws, row: int, headers: list[str]) -> int:
    ws.row_dimensions[row].height = 22
    for c, h in enumerate(headers, 1):
        _cell(ws, row, c, h, bold=True, size=9,
              bg=C_SUBHEADER, fg=C_HEADER_FG, align="center", border=_border())
    return row + 1


def _total_row(ws, row: int, label: str, n_cols: int,
               values: dict[int, float]) -> int:
    """Write a grand-total row. values = {col_index: amount}."""
    ws.row_dimensions[row].height = 22
    ws.merge_cells(f"A{row}:{get_column_letter(n_cols - len(values))}{row}")
    _cell(ws, row, 1, label, bold=True, size=10,
          bg=C_GRAND_BG, fg=C_GRAND_FG, align="right", border=_thick_bottom())
    for col, val in values.items():
        _cell(ws, row, col, val, bold=True, size=10,
              bg=C_GRAND_BG, fg=C_GRAND_FG, align="right",
              num_fmt=INR, border=_thick_bottom())
    return row + 1


# ── Sheet 1 — Cover ───────────────────────────────────────────────────────────

def _write_cover(wb: openpyxl.Workbook, gstr1: dict) -> None:
    ws = wb.create_sheet("Cover")
    ws.sheet_view.showGridLines = False
    _set_widths(ws, [32, 22, 22])

    period  = gstr1.get("period_label") or "—"
    totals  = gstr1.get("totals", {})
    nil     = gstr1.get("nil_rated", {})
    stamp   = datetime.now().strftime("%d %B %Y, %I:%M %p")
    n_warn  = len(gstr1.get("warnings", []))

    r = _title_block(ws, 1, "GSTR-1  —  OUTWARD SUPPLIES RETURN", 3,
                     subtitle=f"Filing Period: {period}  |  Generated: {stamp}")

    sections = [
        ("FILING PERIOD",                  period,                          None),
        (None, None, None),
        ("OUTWARD SUPPLIES AT A GLANCE",   None,                            None),
        ("B2B Invoice Count",              totals.get("b2b_invoice_count", 0), C_GREEN_BG),
        ("B2B Taxable Value (₹)",          totals.get("b2b_taxable",       0), C_GREEN_BG),
        ("B2B CGST (₹)",                   totals.get("b2b_cgst",          0), C_GREEN_BG),
        ("B2B SGST (₹)",                   totals.get("b2b_sgst",          0), C_GREEN_BG),
        ("B2B IGST (₹)",                   totals.get("b2b_igst",          0), C_GREEN_BG),
        ("B2B Cess (₹)",                   totals.get("b2b_cess",          0), C_GREEN_BG),
        (None, None, None),
        ("B2C Large Taxable Value (₹)",    totals.get("b2c_large_taxable", 0), C_SECTION_BG),
        ("B2C Large IGST (₹)",             totals.get("b2c_large_igst",    0), C_SECTION_BG),
        (None, None, None),
        ("Nil-rated Supplies (₹)",         nil.get("total_nil",     0),        C_AMBER_BG),
        ("Exempt Supplies (₹)",            nil.get("total_exempt",  0),        C_AMBER_BG),
        ("Non-GST Supplies (₹)",           nil.get("total_non_gst", 0),        C_AMBER_BG),
        (None, None, None),
        ("TOTAL TAXABLE VALUE (₹)",        totals.get("total_taxable",     0), C_TOTAL_BG),
        ("TOTAL TAX LIABILITY (₹)",        totals.get("total_tax",         0), C_TOTAL_BG),
        (None, None, None),
        ("Data Quality Warnings",          n_warn,                              C_WARN_BG if n_warn else C_GREEN_BG),
    ]

    for label, value, bg in sections:
        ws.row_dimensions[r].height = 20
        if label is None:
            r += 1
            continue
        if value is None:
            ws.merge_cells(f"A{r}:C{r}")
            _cell(ws, r, 1, label, bold=True, size=10,
                  bg=C_HEADER_BG, fg=C_HEADER_FG, border=_border())
        else:
            _cell(ws, r, 1, f"  {label}", size=10, bg=bg, border=_border())
            if isinstance(value, float) and value > 0:
                _cell(ws, r, 2, value, bold=True, size=11, bg=bg,
                      align="right", num_fmt=INR, border=_border())
            else:
                _cell(ws, r, 2, value, bold=True, size=11, bg=bg,
                      align="center", border=_border())
            _cell(ws, r, 3, "", bg=bg, border=_border())
        r += 1

    # Warnings list
    warnings = gstr1.get("warnings", [])
    if warnings:
        r += 1
        ws.row_dimensions[r].height = 22
        ws.merge_cells(f"A{r}:C{r}")
        _cell(ws, r, 1, "DATA QUALITY WARNINGS", bold=True, size=10,
              bg=C_HEADER_BG, fg=C_HEADER_FG, border=_border())
        r += 1
        for i, w in enumerate(warnings, 1):
            ws.row_dimensions[r].height = 40
            _cell(ws, r, 1, i, align="center", size=8,
                  bg=C_WARN_BG, border=_border())
            ws.merge_cells(f"B{r}:C{r}")
            _cell(ws, r, 2, w, size=9, wrap=True,
                  bg=C_WARN_BG, border=_border())
            r += 1


# ── Sheet 2 — B2B (Table 4) ───────────────────────────────────────────────────

def _write_b2b(wb: openpyxl.Workbook, rows: list[dict], period: str) -> None:
    ws = wb.create_sheet("B2B (Table 4)")
    ws.sheet_view.showGridLines = False
    _set_widths(ws, [18, 26, 16, 12, 14, 22, 10, 8, 14, 14, 14, 14, 12, 14])

    r = _title_block(ws, 1,
                     "B2B TAXABLE OUTWARD SUPPLIES",
                     14,
                     subtitle=f"GSTR-1 Table 4  |  Period: {period}  |  Invoice-wise detail for GSTIN-registered buyers",
                     table_ref="Table 4")

    hdrs = [
        "#", "Receiver GSTIN", "Receiver Name",
        "Invoice No.", "Invoice Date", "Place of Supply",
        "Supply Type", "GST Rate %",
        "Taxable Value (₹)", "CGST (₹)", "SGST (₹)", "IGST (₹)",
        "Cess (₹)", "Note Type",
    ]
    r = _col_headers(ws, r, hdrs)
    ws.freeze_panes = f"A{r}"

    tot_taxable = tot_cgst = tot_sgst = tot_igst = tot_cess = 0.0

    for i, row in enumerate(rows, 1):
        alt = C_ROW_ALT if i % 2 == 0 else None
        note_bg = C_AMBER_BG if row["row_type"] != "Regular" else alt
        ws.row_dimensions[r].height = 18

        _cell(ws, r,  1, i, align="center", size=8, bg=alt, border=_border())
        _cell(ws, r,  2, row["receiver_gstin"], bg=alt,  border=_border())
        _cell(ws, r,  3, row["receiver_name"],  bg=alt,  border=_border())
        _cell(ws, r,  4, row["invoice_number"], bg=alt,  border=_border())
        _cell(ws, r,  5, str(row["invoice_date"] or ""), align="center", bg=alt, border=_border())
        _cell(ws, r,  6, row["place_of_supply"], bg=alt, border=_border())
        _cell(ws, r,  7, row["supply_type"].upper(), align="center", bg=alt, border=_border())
        _cell(ws, r,  8, row["gst_rate"], align="center",
              num_fmt="0.0\"%\"", bg=alt, border=_border())
        _cell(ws, r,  9, row["taxable_value"], align="right",
              num_fmt=INR, bg=alt, border=_border())
        _cell(ws, r, 10, row["cgst"] or None, align="right",
              num_fmt=INR, bg=alt, border=_border())
        _cell(ws, r, 11, row["sgst"] or None, align="right",
              num_fmt=INR, bg=alt, border=_border())
        _cell(ws, r, 12, row["igst"] or None, align="right",
              num_fmt=INR, bg=alt, border=_border())
        _cell(ws, r, 13, row["cess"] or None, align="right",
              num_fmt=INR, bg=alt, border=_border())
        _cell(ws, r, 14, row["row_type"], align="center",
              bg=note_bg, border=_border())

        tot_taxable += row["taxable_value"]
        tot_cgst    += row["cgst"]
        tot_sgst    += row["sgst"]
        tot_igst    += row["igst"]
        tot_cess    += row["cess"]
        r += 1

    _total_row(ws, r, "TOTAL", 14, {
        9: round(tot_taxable, 2),
        10: round(tot_cgst,   2),
        11: round(tot_sgst,   2),
        12: round(tot_igst,   2),
        13: round(tot_cess,   2),
    })


# ── Sheet 3 — B2C Large (Table 7) ────────────────────────────────────────────

def _write_b2c_large(wb: openpyxl.Workbook, rows: list[dict], period: str) -> None:
    ws = wb.create_sheet("B2C Large (Table 7)")
    ws.sheet_view.showGridLines = False
    _set_widths(ws, [4, 28, 12, 16, 14, 12, 12])

    r = _title_block(ws, 1,
                     "B2C LARGE — INTER-STATE OUTWARD SUPPLIES",
                     7,
                     subtitle=f"GSTR-1 Table 7  |  Period: {period}  |  Inter-state invoices > ₹2.5 lakh without buyer GSTIN",
                     table_ref="Table 7")

    hdrs = ["#", "Place of Supply", "GST Rate %",
            "Taxable Value (₹)", "IGST (₹)", "Cess (₹)", "Invoice Count"]
    r = _col_headers(ws, r, hdrs)
    ws.freeze_panes = f"A{r}"

    tot_taxable = tot_igst = tot_cess = 0
    for i, row in enumerate(rows, 1):
        alt = C_ROW_ALT if i % 2 == 0 else None
        ws.row_dimensions[r].height = 18
        _cell(ws, r, 1, i, align="center", size=8, bg=alt, border=_border())
        _cell(ws, r, 2, row["place_of_supply"], bg=alt, border=_border())
        _cell(ws, r, 3, row["gst_rate"], align="center",
              num_fmt="0.0\"%\"", bg=alt, border=_border())
        _cell(ws, r, 4, row["taxable_value"], align="right",
              num_fmt=INR, bg=alt, border=_border())
        _cell(ws, r, 5, row["igst"] or None, align="right",
              num_fmt=INR, bg=alt, border=_border())
        _cell(ws, r, 6, row["cess"] or None, align="right",
              num_fmt=INR, bg=alt, border=_border())
        _cell(ws, r, 7, row["invoice_count"], align="center", bg=alt, border=_border())
        tot_taxable += row["taxable_value"]
        tot_igst    += row["igst"]
        tot_cess    += row["cess"]
        r += 1

    _total_row(ws, r, "TOTAL", 7, {
        4: round(tot_taxable, 2),
        5: round(tot_igst,    2),
        6: round(tot_cess,    2),
    })


# ── Sheet 4 — Nil & Exempt (Table 8) ─────────────────────────────────────────

def _write_nil_exempt(wb: openpyxl.Workbook, nil: dict, period: str) -> None:
    ws = wb.create_sheet("Nil & Exempt (Table 8)")
    ws.sheet_view.showGridLines = False
    _set_widths(ws, [28, 20, 20, 20])

    r = _title_block(ws, 1,
                     "NIL-RATED, EXEMPT & NON-GST OUTWARD SUPPLIES",
                     4,
                     subtitle=f"GSTR-1 Table 8  |  Period: {period}",
                     table_ref="Table 8")

    r = _col_headers(ws, r, ["Category", "Inter-state (₹)", "Intra-state (₹)", "Total (₹)"])

    data_rows = [
        ("Nil-Rated Supplies",  nil["nil_rated_b2b"],  nil["nil_rated_b2c"],  nil["total_nil"]),
        ("Exempt Supplies",     nil["exempt_b2b"],     nil["exempt_b2c"],     nil["total_exempt"]),
        ("Non-GST Supplies",    nil["non_gst_b2b"],    nil["non_gst_b2c"],    nil["total_non_gst"]),
    ]

    for i, (label, b2b, b2c, total) in enumerate(data_rows, 1):
        alt = C_ROW_ALT if i % 2 == 0 else None
        ws.row_dimensions[r].height = 22
        _cell(ws, r, 1, label, bold=True, bg=alt, border=_border())
        _cell(ws, r, 2, b2b   or None, align="right", num_fmt=INR, bg=alt, border=_border())
        _cell(ws, r, 3, b2c   or None, align="right", num_fmt=INR, bg=alt, border=_border())
        _cell(ws, r, 4, total or None, align="right", num_fmt=INR, bg=C_TOTAL_BG, border=_border())
        r += 1

    # Grand total
    ws.row_dimensions[r].height = 22
    grand = round(nil["total_nil"] + nil["total_exempt"] + nil["total_non_gst"], 2)
    grand_b2b = round(nil["nil_rated_b2b"] + nil["exempt_b2b"] + nil["non_gst_b2b"], 2)
    grand_b2c = round(nil["nil_rated_b2c"] + nil["exempt_b2c"] + nil["non_gst_b2c"], 2)
    _cell(ws, r, 1, "TOTAL", bold=True, size=10,
          bg=C_GRAND_BG, fg=C_GRAND_FG, align="right", border=_thick_bottom())
    _cell(ws, r, 2, grand_b2b or None, bold=True, size=10,
          bg=C_GRAND_BG, fg=C_GRAND_FG, align="right", num_fmt=INR, border=_thick_bottom())
    _cell(ws, r, 3, grand_b2c or None, bold=True, size=10,
          bg=C_GRAND_BG, fg=C_GRAND_FG, align="right", num_fmt=INR, border=_thick_bottom())
    _cell(ws, r, 4, grand or None, bold=True, size=10,
          bg=C_GRAND_BG, fg=C_GRAND_FG, align="right", num_fmt=INR, border=_thick_bottom())
    r += 1

    # Notes
    r += 1
    notes = [
        "Nil-Rated : Supplies attracting 0% GST (e.g. fresh vegetables, grains, books).",
        "Exempt    : Supplies exempted under GST Act (e.g. healthcare, education services).",
        "Non-GST   : Supplies outside GST purview (e.g. petrol, alcohol, electricity).",
        "B2B column = inter-state supplies to GSTIN-registered buyers.",
        "B2C column = all other nil/exempt/non-GST supplies.",
    ]
    for note in notes:
        ws.row_dimensions[r].height = 16
        ws.merge_cells(f"A{r}:D{r}")
        _cell(ws, r, 1, note, size=8, italic=True, fg="595959")
        r += 1


# ── Sheet 5 — HSN Summary (Table 12) ─────────────────────────────────────────

def _write_hsn_summary(wb: openpyxl.Workbook, rows: list[dict], period: str) -> None:
    ws = wb.create_sheet("HSN Summary (Table 12)")
    ws.sheet_view.showGridLines = False
    _set_widths(ws, [4, 12, 30, 8, 10, 14, 16, 14, 14, 14, 14, 12])

    r = _title_block(ws, 1,
                     "HSN/SAC-WISE SUMMARY OF OUTWARD SUPPLIES",
                     12,
                     subtitle=f"GSTR-1 Table 12  |  Period: {period}  |  Grouped by HSN code, UOM, and GST rate",
                     table_ref="Table 12")

    hdrs = [
        "#", "HSN / SAC", "Description", "UOM", "GST Rate %",
        "Total Qty", "Total Value (₹)", "Taxable Value (₹)",
        "CGST (₹)", "SGST (₹)", "IGST (₹)", "Cess (₹)",
    ]
    r = _col_headers(ws, r, hdrs)
    ws.freeze_panes = f"A{r}"

    tot_val = tot_tax = tot_cgst = tot_sgst = tot_igst = tot_cess = 0.0

    for i, row in enumerate(rows, 1):
        alt    = C_ROW_ALT if i % 2 == 0 else None
        is_unk = row["hsn_sac"] == "UNKNOWN"
        hsn_bg = C_RED_BG if is_unk else alt
        ws.row_dimensions[r].height = 18

        _cell(ws, r,  1, i, align="center", size=8, bg=alt,  border=_border())
        _cell(ws, r,  2, row["hsn_sac"],    bg=hsn_bg, border=_border(),
              bold=is_unk, fg="CC0000" if is_unk else "000000")
        _cell(ws, r,  3, row["description"], bg=alt, border=_border())
        _cell(ws, r,  4, row["uom"], align="center", bg=alt, border=_border())
        _cell(ws, r,  5, row["gst_rate"], align="center",
              num_fmt="0.0\"%\"", bg=alt, border=_border())
        _cell(ws, r,  6, row["total_quantity"] or None, align="right",
              num_fmt="#,##0.000", bg=alt, border=_border())
        _cell(ws, r,  7, row["total_value"]   or None, align="right",
              num_fmt=INR, bg=alt, border=_border())
        _cell(ws, r,  8, row["taxable_value"] or None, align="right",
              num_fmt=INR, bg=alt, border=_border())
        _cell(ws, r,  9, row["cgst"] or None, align="right",
              num_fmt=INR, bg=alt, border=_border())
        _cell(ws, r, 10, row["sgst"] or None, align="right",
              num_fmt=INR, bg=alt, border=_border())
        _cell(ws, r, 11, row["igst"] or None, align="right",
              num_fmt=INR, bg=alt, border=_border())
        _cell(ws, r, 12, row["cess"] or None, align="right",
              num_fmt=INR, bg=alt, border=_border())

        tot_val  += row["total_value"]
        tot_tax  += row["taxable_value"]
        tot_cgst += row["cgst"]
        tot_sgst += row["sgst"]
        tot_igst += row["igst"]
        tot_cess += row["cess"]
        r += 1

    _total_row(ws, r, "TOTAL", 12, {
        7:  round(tot_val,  2),
        8:  round(tot_tax,  2),
        9:  round(tot_cgst, 2),
        10: round(tot_sgst, 2),
        11: round(tot_igst, 2),
        12: round(tot_cess, 2),
    })


def write_gstr1_xlsx(
    gstr1:       dict,
    output_path: Path | str,
) -> None:
    """
    Write a GSTR-1 workbook from the dict returned by build_gstr1().

    Args:
        gstr1       : dict from gstr1_builder.build_gstr1()
        output_path : destination .xlsx file path
    """
    path   = Path(output_path)
    wb     = openpyxl.Workbook()
    wb.remove(wb.active)

    period = gstr1.get("period_label") or "—"
    b2b    = gstr1.get("b2b",         [])
    b2c    = gstr1.get("b2c_large",   [])
    nil    = gstr1.get("nil_rated",   {})
    hsn    = gstr1.get("hsn_summary", [])

    _write_cover(wb, gstr1)
    _write_b2b(wb, b2b, period)
    _write_b2c_large(wb, b2c, period)
    _write_nil_exempt(wb, nil, period)
    _write_hsn_summary(wb, hsn, period)

    wb.active = wb["Cover"]
    wb.save(path)

    t = gstr1.get("totals", {})
    print(
        f"  GSTR-1 saved → {path}\n"
        f"    Period: {period}  |  B2B: {len(b2b)} invoices  "
        f"|  HSN rows: {len(hsn)}  |  Warnings: {len(gstr1.get('warnings', []))}\n"
        f"    Total taxable: ₹{t.get('total_taxable', 0):,.2f}  "
        f"|  Total tax: ₹{t.get('total_tax', 0):,.2f}"
    )
