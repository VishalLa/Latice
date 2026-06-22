"""
journal_xlsx.py — Journal Book Export to XLSX

Writes all JournalEntry objects produced by build_ledger() / to_journal_entries()
into a formatted Excel workbook.

Sheets produced:
  1. Journal Book  — complete audit trail, one row per entry-line
  2. Summary       — entry count & totals grouped by voucher type
"""

from __future__ import annotations

from collections import defaultdict
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from schema import DrCr, JournalEntry

# ── Colour palette (matches excel.py / Reports.py) ──────────────────────────
C_HEADER_BG  = "1E3A5F"
C_HEADER_FG  = "FFFFFF"
C_SUBHEADER  = "2E6DA4"
C_ROW_ALT    = "EBF3FB"
C_TOTAL_BG   = "FFF2CC"
C_GRAND_BG   = "1E3A5F"
C_GRAND_FG   = "FFFFFF"
C_SECTION_BG = "D9E1F2"
C_BORDER     = "BFBFBF"
C_DEBIT_BG   = "FFF0E0"
C_CREDIT_BG  = "E0F0FF"

INR = "\u20b9#,##0.00"


# ── Helpers ──────────────────────────────────────────────────────────────────

def _border(style="thin", color=C_BORDER):
    s = Side(style=style, color=color)
    return Border(left=s, right=s, top=s, bottom=s)


def _thick_bottom():
    t = Side(style="thin",   color=C_BORDER)
    m = Side(style="medium", color="000000")
    return Border(left=t, right=t, top=t, bottom=m)


def _cell(ws, row, col, value=None, *, bold=False, size=9, bg=None,
          fg="000000", align="left", num_fmt=None, italic=False,
          border=None, wrap=False):
    c = ws.cell(row=row, column=col, value=value)
    c.font      = Font(name="Arial", bold=bold, size=size, color=fg, italic=italic)
    c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
    if bg:
        c.fill = PatternFill("solid", fgColor=bg)
    if num_fmt:
        c.number_format = num_fmt
    if border:
        c.border = border
    return c


def _title(ws, row: int, text: str, n_cols: int, subtitle: str = "") -> int:
    ws.row_dimensions[row].height = 32
    ws.merge_cells(f"A{row}:{get_column_letter(n_cols)}{row}")
    _cell(ws, row, 1, text, bold=True, size=14,
          bg=C_HEADER_BG, fg=C_HEADER_FG, align="center")
    row += 1
    if subtitle:
        ws.row_dimensions[row].height = 18
        ws.merge_cells(f"A{row}:{get_column_letter(n_cols)}{row}")
        _cell(ws, row, 1, subtitle, size=9,
              bg=C_SUBHEADER, fg=C_HEADER_FG, align="center", italic=True)
        row += 1
    return row + 1          # blank gap row


# ── Sheet 1 — Journal Book ───────────────────────────────────────────────────

def _write_journal_book(wb: openpyxl.Workbook, entries: list[JournalEntry]) -> None:
    ws = wb.create_sheet("Journal Book")
    ws.sheet_view.showGridLines = False

    col_widths = [5, 12, 18, 36, 28, 12, 16, 16]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    stamp = datetime.now().strftime("%d %B %Y, %I:%M %p")
    r = _title(
        ws, 1,
        "JOURNAL BOOK (AUDIT TRAIL)",
        8,
        f"Generated: {stamp}  |  Total entries: {len(entries)}",
    )

    hdrs = ["#", "Date", "Voucher Type", "Narration",
            "Account", "Dr/Cr", "Debit (₹)", "Credit (₹)"]
    ws.row_dimensions[r].height = 24
    for c, h in enumerate(hdrs, 1):
        _cell(ws, r, c, h, bold=True, size=9,
              bg=C_SUBHEADER, fg=C_HEADER_FG, align="center", border=_border())
    r += 1
    ws.freeze_panes = f"A{r}"

    sorted_entries = sorted(entries, key=lambda e: e.date)
    grand_dr = grand_cr = 0.0

    for idx, entry in enumerate(sorted_entries, 1):
        lines = entry.lines or []
        for li, line in enumerate(lines):
            alt = C_ROW_ALT if idx % 2 == 0 else None
            dr_val = line.amount if line.dr_cr == DrCr.DEBIT  else None
            cr_val = line.amount if line.dr_cr == DrCr.CREDIT else None
            if dr_val:
                grand_dr += dr_val
            if cr_val:
                grand_cr += cr_val

            ws.row_dimensions[r].height = 18
            if li == 0:
                _cell(ws, r, 1, idx, align="center", bg=alt, border=_border())
                _cell(ws, r, 2, entry.date.strftime("%d-%m-%Y"),
                      align="center", bg=alt, border=_border())
                _cell(ws, r, 3, entry.voucher_type, bg=alt, border=_border())
                _cell(ws, r, 4, entry.narration, bg=alt, wrap=True, border=_border())
            else:
                for col in range(1, 5):
                    _cell(ws, r, col, "", bg=alt, border=_border())

            indent = "" if line.dr_cr == DrCr.DEBIT else "    "
            _cell(ws, r, 5, f"{indent}{line.account.name}", bg=alt, border=_border())
            _cell(ws, r, 6, line.dr_cr.value, align="center",
                  bg=C_DEBIT_BG if line.dr_cr == DrCr.DEBIT else C_CREDIT_BG,
                  border=_border())
            _cell(ws, r, 7, dr_val, align="right", num_fmt=INR,
                  bg=alt or C_DEBIT_BG, border=_border())
            _cell(ws, r, 8, cr_val, align="right", num_fmt=INR,
                  bg=alt or C_CREDIT_BG, border=_border())
            r += 1

        # Entry sub-total separator
        ws.row_dimensions[r].height = 4
        ws.merge_cells(f"A{r}:H{r}")
        _cell(ws, r, 1, "", bg=C_SECTION_BG)
        r += 1

    # Grand total
    ws.row_dimensions[r].height = 24
    ws.merge_cells(f"A{r}:F{r}")
    _cell(ws, r, 1, "GRAND TOTAL", bold=True, size=11,
          bg=C_GRAND_BG, fg=C_GRAND_FG, align="right", border=_thick_bottom())
    _cell(ws, r, 7, grand_dr, bold=True, size=11,
          bg=C_GRAND_BG, fg=C_GRAND_FG, align="right",
          num_fmt=INR, border=_thick_bottom())
    _cell(ws, r, 8, grand_cr, bold=True, size=11,
          bg=C_GRAND_BG, fg=C_GRAND_FG, align="right",
          num_fmt=INR, border=_thick_bottom())


# ── Sheet 2 — Summary ────────────────────────────────────────────────────────

def _write_summary(wb: openpyxl.Workbook, entries: list[JournalEntry]) -> None:
    ws = wb.create_sheet("Summary")
    ws.sheet_view.showGridLines = False

    col_widths = [28, 10, 18, 18]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    r = _title(ws, 1, "JOURNAL ENTRY SUMMARY", 4,
               "Entry counts and totals grouped by voucher type")

    hdrs = ["Voucher Type", "Entries", "Total Debit (₹)", "Total Credit (₹)"]
    ws.row_dimensions[r].height = 22
    for c, h in enumerate(hdrs, 1):
        _cell(ws, r, c, h, bold=True, size=9,
              bg=C_SUBHEADER, fg=C_HEADER_FG, align="center", border=_border())
    r += 1

    by_type: dict[str, list[JournalEntry]] = defaultdict(list)
    for entry in entries:
        by_type[entry.voucher_type].append(entry)

    grand_count = grand_dr = grand_cr = 0
    for row_i, vtype in enumerate(sorted(by_type), 1):
        group = by_type[vtype]
        count = len(group)
        total_dr = sum(
            l.amount for e in group for l in e.lines if l.dr_cr == DrCr.DEBIT
        )
        total_cr = sum(
            l.amount for e in group for l in e.lines if l.dr_cr == DrCr.CREDIT
        )
        grand_count += count
        grand_dr    += total_dr
        grand_cr    += total_cr

        alt = C_ROW_ALT if row_i % 2 == 0 else None
        ws.row_dimensions[r].height = 18
        _cell(ws, r, 1, vtype, bg=alt, border=_border())
        _cell(ws, r, 2, count, align="center", bg=alt, border=_border())
        _cell(ws, r, 3, total_dr or None, align="right",
              num_fmt=INR, bg=alt, border=_border())
        _cell(ws, r, 4, total_cr or None, align="right",
              num_fmt=INR, bg=alt, border=_border())
        r += 1

    ws.row_dimensions[r].height = 22
    _cell(ws, r, 1, "TOTAL", bold=True, size=10,
          bg=C_GRAND_BG, fg=C_GRAND_FG, align="right", border=_thick_bottom())
    _cell(ws, r, 2, grand_count, bold=True, size=10,
          bg=C_GRAND_BG, fg=C_GRAND_FG, align="center", border=_thick_bottom())
    _cell(ws, r, 3, grand_dr or None, bold=True, size=10,
          bg=C_GRAND_BG, fg=C_GRAND_FG, align="right",
          num_fmt=INR, border=_thick_bottom())
    _cell(ws, r, 4, grand_cr or None, bold=True, size=10,
          bg=C_GRAND_BG, fg=C_GRAND_FG, align="right",
          num_fmt=INR, border=_thick_bottom())


# ── Public API ───────────────────────────────────────────────────────────────

def write_journal_xlsx(
    entries:     list[JournalEntry],
    output_path: Path | str,
) -> None:
    """
    Write all journal entries to an Excel workbook.

    Args:
        entries     : list[JournalEntry] — from build_ledger() or to_journal_entries()
        output_path : where to save the .xlsx file
    """
    path = Path(output_path)
    wb   = openpyxl.Workbook()
    wb.remove(wb.active)

    _write_journal_book(wb, entries)
    _write_summary(wb, entries)

    wb.active = wb["Journal Book"]
    wb.save(path)
    print(f"  Journal book saved → {path}  ({len(entries)} entries, 2 sheets)")
    