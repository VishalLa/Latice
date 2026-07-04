from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional, Sequence, Union

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from schema import LedgerFormat, BankStatement
from matcher.confidence import confidence_bucket

# ── Colour palette (matches existing report styling) ────────────────────────
C_HEADER_BG  = "1E3A5F"
C_HEADER_FG  = "FFFFFF"
C_SUBHEADER  = "2E6DA4"
C_ROW_ALT    = "EBF3FB"
C_SECTION_BG = "D9E1F2"
C_BORDER     = "BFBFBF"
C_GREEN_BG   = "E2EFDA"
C_AMBER_BG   = "FFF2CC"
C_RED_BG     = "FFCCCC"
C_ORANGE_BG  = "FCE4D6"

INR = "\u20b9#,##0.00"

_CONF_COLOURS = {
    "High":   C_GREEN_BG,
    "Medium": C_AMBER_BG,
    "Low":    C_ORANGE_BG,
}


# ── Generic cell / layout helpers ────────────────────────────────────────────

def _border(style="thin", color=C_BORDER):
    s = Side(style=style, color=color)
    return Border(left=s, right=s, top=s, bottom=s)


def _cell(ws, row, col, value=None, *, bold=False, size=9, bg=None,
          fg="000000", align="left", num_fmt=None, italic=False,
          border=None, wrap=False):
    c = ws.cell(row=row, column=col, value=value)
    c.font      = Font(name="Arial", bold=bold, size=size, color=fg, italic=italic)
    c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
    if bg:
        c.fill = PatternFill("solid", fgColor=bg)
    if num_fmt:
        c.number_format = num_fmt
    if border:
        c.border = border
    return c


def _title(ws, row: int, text: str, n_cols: int, subtitle: str = "") -> int:
    ws.row_dimensions[row].height = 32
    ws.merge_cells(f"A{row}:{get_column_letter(n_cols)}{row}")
    _cell(ws, row, 1, text, bold=True, size=14,
          bg=C_HEADER_BG, fg=C_HEADER_FG, align="center")
    row += 1
    if subtitle:
        ws.row_dimensions[row].height = 18
        ws.merge_cells(f"A{row}:{get_column_letter(n_cols)}{row}")
        _cell(ws, row, 1, subtitle, size=9,
              bg=C_SUBHEADER, fg=C_HEADER_FG, align="center", italic=True)
        row += 1
    return row + 1


def _col_headers(ws, row: int, headers: List[str],
                  bg=C_SUBHEADER, fg=C_HEADER_FG) -> int:
    ws.row_dimensions[row].height = 22
    for c, h in enumerate(headers, 1):
        _cell(ws, row, c, h, bold=True, size=9,
              bg=bg, fg=fg, align="center", border=_border())
    return row + 1


def _section_row(ws, row: int, text: str, n_cols: int) -> int:
    ws.row_dimensions[row].height = 20
    ws.merge_cells(f"A{row}:{get_column_letter(n_cols)}{row}")
    _cell(ws, row, 1, text, bold=True, size=10,
          bg=C_HEADER_BG, fg=C_HEADER_FG, border=_border())
    return row + 1


# ── Match-index builders ─────────────────────────────────────────────────────
# Every match dict, regardless of which phase produced it, may carry the
# matched id(s) as either a single "ledger_id" / "bank_id" string, an
# " & "-joined composite string (many-to-one fuzzy matches), or an
# "ledger_ids": [{"ledger_id": ...}, ...] list (AI many-to-one matches).

def _split_ids(raw) -> List[str]:
    if raw is None:
        return []
    if isinstance(raw, list):
        out = []
        for item in raw:
            if isinstance(item, dict):
                v = item.get("ledger_id") or item.get("bank_id")
            else:
                v = item
            if v not in (None, ""):
                out.append(str(v))
        return out
    s = str(raw).strip()
    if not s or s == "—":
        return []
    return [p.strip() for p in s.split("&") if p.strip()]


def _resolve_ledger_ids(m: dict) -> List[str]:
    if m.get("ledger_ids"):
        return _split_ids(m["ledger_ids"])
    return _split_ids(m.get("ledger_id"))


def _resolve_bank_ids(m: dict) -> List[str]:
    return _split_ids(m.get("bank_id"))


PHASE_LABELS = {
    "exact":  "Exact",
    "fuzzy":  "Fuzzy",
    "memory": "Memory",
    "ai":     "AI",
}


def _all_matches(recon_result: dict) -> List[dict]:
    """Flatten every phase's matches into one list, each tagged with phase."""
    out: List[dict] = []
    for key, phase in (
        ("EXACT_MATCHES",  "exact"),
        ("FUZZY_MATCHES",  "fuzzy"),
        ("MEMORY_MATCHES", "memory"),
        ("AI_MATCHES",     "ai"),
    ):
        for m in recon_result.get(key, []):
            mm = dict(m)
            mm.setdefault("match_phase", phase)
            out.append(mm)
    return out


def _match_explanation(m: dict, phase: str) -> str:
    if phase == "exact":
        parts = ["Matched on identical transaction date and amount."]
        if m.get("reference_matched"):
            parts.append("Confirmed by matching reference/cheque number.")
        else:
            parts.append("No reference number confirmation was available; "
                          "matched on amount + date only.")
        return " ".join(parts)

    if phase == "memory":
        return m.get("details") or (
            "Counterparty pattern recognized from a prior reconciliation run."
        )

    if phase == "ai":
        reasoning = m.get("reasoning") or m.get("details") or ""
        if m.get("ledger_ids"):
            prefix = "AI matched several ledger entries summing to one bank line. "
        else:
            prefix = "AI semantic match (no exact date/amount/reference overlap). "
        return (prefix + reasoning).strip()

    # fuzzy
    return m.get("details") or m.get("adjustment_type") or "Fuzzy rule match."


def _match_adjustment_type(m: dict, phase: str) -> str:
    if phase == "exact":
        return "Exact Match"
    if phase == "ai":
        return "AI Many-to-One Match" if m.get("ledger_ids") else "AI Semantic Match"
    return m.get("adjustment_type") or "—"


def _build_match_index(all_matches: list[dict]) -> tuple[dict[str, dict], dict[str, dict]]:
    """Map ledger_id -> match summary, bank_id -> match summary, for the
    Ledger Entries / Bank Entries sheets."""
    ledger_idx: Dict[str, dict] = {}
    bank_idx:   Dict[str, dict] = {}
    for m in all_matches:
        phase = m.get("match_phase", "")
        lids = _resolve_ledger_ids(m)
        bids = _resolve_bank_ids(m)
        info = {
            "phase":     PHASE_LABELS.get(phase, phase.title() or "—"),
            "ledger_id": " & ".join(lids) or "—",
            "bank_id":   " & ".join(bids) or "—",
        }
        for lid in lids:
            ledger_idx[lid] = info
        for bid in bids:
            bank_idx[bid] = info
    return ledger_idx, bank_idx


# ── Sheet 1 — Ledger Entries ─────────────────────────────────────────────────

def _write_ledger_entries(wb, gl_records: Sequence[LedgerFormat],
                           ledger_idx: Dict[str, dict]) -> None:
    ws = wb.create_sheet("Ledger Entries")
    ws.sheet_view.showGridLines = False
    for i, w in enumerate([12, 12, 26, 20, 16, 12, 14, 14, 10, 12, 14, 12], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    r = _title(ws, 1, "LEDGER ENTRIES", 12,
               f"{len(gl_records)} ledger record(s) loaded")
    r = _col_headers(ws, r, [
        "Ledger ID", "Date", "Account Name", "Vendor", "Voucher Type",
        "Reference", "Debit (₹)", "Credit (₹)", "Source",
        "Match Status", "Matched Via", "Matched Bank ID",
    ])
    ws.freeze_panes = f"A{r}"

    for i, rec in enumerate(gl_records, 1):
        lid  = str(rec.ledger_id or "")
        info = ledger_idx.get(lid)
        matched = info is not None
        alt = C_ROW_ALT if i % 2 == 0 else None
        bg  = C_GREEN_BG if matched else (alt or C_RED_BG)

        ws.row_dimensions[r].height = 18
        _cell(ws, r, 1, lid,                                   bg=bg, border=_border())
        _cell(ws, r, 2, rec.transaction_date or rec.transaction_date_raw or "",
              align="center", bg=bg, border=_border())
        _cell(ws, r, 3, rec.account_name or "",                bg=bg, border=_border())
        _cell(ws, r, 4, rec.vendor_name or "",                 bg=bg, border=_border())
        _cell(ws, r, 5, rec.voucher_type or "", size=8,        bg=bg, border=_border())
        _cell(ws, r, 6, rec.reference_id or "", align="center", bg=bg, border=_border())
        _cell(ws, r, 7, rec.debit_amount or None,  align="right", num_fmt=INR, bg=bg, border=_border())
        _cell(ws, r, 8, rec.credit_amount or None, align="right", num_fmt=INR, bg=bg, border=_border())
        _cell(ws, r, 9, rec.source.value if rec.source else "", size=8, bg=bg, border=_border())
        _cell(ws, r, 10, "Matched" if matched else "Unmatched",
              align="center", bold=True, bg=bg, border=_border())
        _cell(ws, r, 11, info["phase"]     if info else "—", align="center", bg=bg, border=_border())
        _cell(ws, r, 12, info["bank_id"]   if info else "—", align="center", bg=bg, border=_border())
        r += 1


# ── Sheet 2 — Bank Entries ───────────────────────────────────────────────────

def _write_bank_entries(wb, bank_records: Sequence[BankStatement],
                         bank_idx: Dict[str, dict]) -> None:
    ws = wb.create_sheet("Bank Entries")
    ws.sheet_view.showGridLines = False
    for i, w in enumerate([10, 12, 32, 16, 14, 14, 14, 14, 12, 14, 14], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    r = _title(ws, 1, "BANK ENTRIES", 11,
               f"{len(bank_records)} bank statement row(s) loaded")
    r = _col_headers(ws, r, [
        "Row Index", "Date", "Narration", "Txn ID",
        "Debit (₹)", "Credit (₹)", "Balance (₹)", "Bank",
        "Match Status", "Matched Via", "Matched Ledger ID",
    ])
    ws.freeze_panes = f"A{r}"

    for i, rec in enumerate(bank_records, 1):
        bid  = str(rec.row_index)
        info = bank_idx.get(bid)
        matched = info is not None
        alt = C_ROW_ALT if i % 2 == 0 else None
        bg  = C_GREEN_BG if matched else (alt or C_RED_BG)

        ws.row_dimensions[r].height = 18
        _cell(ws, r, 1, bid, align="center", bg=bg, border=_border())
        _cell(ws, r, 2, rec.date or rec.date_raw or "", align="center", bg=bg, border=_border())
        _cell(ws, r, 3, rec.narration or "", bg=bg, border=_border())
        _cell(ws, r, 4, rec.txn_id or "", align="center", bg=bg, border=_border())
        _cell(ws, r, 5, rec.debit  or None, align="right", num_fmt=INR, bg=bg, border=_border())
        _cell(ws, r, 6, rec.credit or None, align="right", num_fmt=INR, bg=bg, border=_border())
        _cell(ws, r, 7, rec.balance, align="right", num_fmt=INR, bg=bg, border=_border())
        _cell(ws, r, 8, rec.bank_name or "", size=8, bg=bg, border=_border())
        _cell(ws, r, 9, "Matched" if matched else "Unmatched",
              align="center", bold=True, bg=bg, border=_border())
        _cell(ws, r, 10, info["phase"]      if info else "—", align="center", bg=bg, border=_border())
        _cell(ws, r, 11, info["ledger_id"]  if info else "—", align="center", bg=bg, border=_border())
        r += 1


# ── Sheet 3 — Matched & Unmatched (matched on top, unmatched below) ─────────

def _write_matched_unmatched(
    wb,
    recon_result: dict,
    gl_by_id: Dict[str, LedgerFormat],
    bank_by_id: Dict[str, BankStatement],
) -> None:
    ws = wb.create_sheet("Matched & Unmatched")
    ws.sheet_view.showGridLines = False
    widths = [5, 12, 10, 14, 14, 12, 22, 14, 26, 14, 20, 12, 55]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    matches = _all_matches(recon_result)
    unreconciled = recon_result.get("UNRECONCILED_ITEMS", {})
    ledger_unmatched: List[LedgerFormat] = unreconciled.get("ledger", [])
    bank_unmatched:   List[BankStatement] = unreconciled.get("bank", [])

    total = len(matches) + len(ledger_unmatched) + len(bank_unmatched)
    r = _title(ws, 1, "MATCHED & UNMATCHED — FULL RECONCILIATION DETAIL", 13,
               f"{len(matches)} matched pair(s) on top, "
               f"{len(ledger_unmatched) + len(bank_unmatched)} unreconciled item(s) below "
               f"({total} total)")

    headers = [
        "#", "Status", "Phase", "Ledger ID", "Bank ID", "Date",
        "Ledger Account", "Ledger Amt (₹)", "Bank Narration", "Bank Amt (₹)",
        "Adjustment / Reason", "Confidence", "Explanation",
    ]

    # ── MATCHED block ────────────────────────────────────────────────────────
    r = _section_row(ws, r, f"  MATCHED ({len(matches)} items)", 13)
    r = _col_headers(ws, r, headers)
    ws.freeze_panes = f"A{r}"

    idx = 0
    for m in matches:
        idx += 1
        phase = m.get("match_phase", "")
        phase_label = PHASE_LABELS.get(phase, phase.title() or "—")

        lids = _resolve_ledger_ids(m)
        bids = _resolve_bank_ids(m)
        l_recs = [gl_by_id[i] for i in lids if i in gl_by_id]
        b_recs = [bank_by_id[i] for i in bids if i in bank_by_id]

        l_date   = (l_recs[0].transaction_date if l_recs else m.get("date")) or ""
        l_acct   = "; ".join(x.account_name for x in l_recs) or "—"
        l_amt    = sum((x.debit_amount or x.credit_amount) for x in l_recs) or m.get("amount")
        b_narr   = "; ".join(x.narration for x in b_recs) or "—"
        b_amt    = sum((x.debit or x.credit) for x in b_recs) or m.get("amount")

        conf_num = m.get("confidence_numeric")
        conf     = confidence_bucket(conf_num) if conf_num is not None else "High"
        bg       = _CONF_COLOURS.get(conf, C_GREEN_BG)

        ws.row_dimensions[r].height = 20
        _cell(ws, r, 1, idx, align="center", size=8, bg=bg, border=_border())
        _cell(ws, r, 2, "Matched", align="center", bold=True, bg=bg, border=_border())
        _cell(ws, r, 3, phase_label, align="center", bg=bg, border=_border())
        _cell(ws, r, 4, " & ".join(lids) or "—", bg=bg, border=_border())
        _cell(ws, r, 5, " & ".join(bids) or "—", bg=bg, border=_border())
        _cell(ws, r, 6, str(l_date), align="center", bg=bg, border=_border())
        _cell(ws, r, 7, l_acct, bg=bg, border=_border())
        _cell(ws, r, 8, l_amt, align="right", num_fmt=INR, bg=bg, border=_border())
        _cell(ws, r, 9, b_narr, bg=bg, border=_border())
        _cell(ws, r, 10, b_amt, align="right", num_fmt=INR, bg=bg, border=_border())
        _cell(ws, r, 11, _match_adjustment_type(m, phase), bg=bg, border=_border())
        _cell(ws, r, 12, conf, align="center", bold=True, bg=bg, border=_border())
        _cell(ws, r, 13, _match_explanation(m, phase), size=8, italic=True,
              wrap=True, bg=bg, border=_border())
        r += 1

    r += 1  # blank spacer row

    # ── UNMATCHED block ──────────────────────────────────────────────────────
    total_unmatched = len(ledger_unmatched) + len(bank_unmatched)
    r = _section_row(ws, r, f"  UNMATCHED ({total_unmatched} items)", 13)
    r = _col_headers(ws, r, headers)

    idx = 0
    for rec in ledger_unmatched:
        idx += 1
        amt = rec.debit_amount or rec.credit_amount
        ws.row_dimensions[r].height = 20
        _cell(ws, r, 1, idx, align="center", size=8, bg=C_RED_BG, border=_border())
        _cell(ws, r, 2, "Unmatched", align="center", bold=True, bg=C_RED_BG, border=_border())
        _cell(ws, r, 3, "Ledger", align="center", bg=C_RED_BG, border=_border())
        _cell(ws, r, 4, rec.ledger_id or "—", bg=C_RED_BG, border=_border())
        _cell(ws, r, 5, "—", align="center", bg=C_RED_BG, border=_border())
        _cell(ws, r, 6, rec.transaction_date or rec.transaction_date_raw or "",
              align="center", bg=C_RED_BG, border=_border())
        _cell(ws, r, 7, rec.account_name or "", bg=C_RED_BG, border=_border())
        _cell(ws, r, 8, amt, align="right", num_fmt=INR, bg=C_RED_BG, border=_border())
        _cell(ws, r, 9, "—", bg=C_RED_BG, border=_border())
        _cell(ws, r, 10, None, align="right", num_fmt=INR, bg=C_RED_BG, border=_border())
        _cell(ws, r, 11, "No Match Found", bg=C_RED_BG, border=_border())
        _cell(ws, r, 12, "—", align="center", bg=C_RED_BG, border=_border())
        _cell(ws, r, 13,
              "No corresponding bank row was found in the Exact, Fuzzy, Memory or "
              "AI matching phases within the configured tolerances. Requires manual review.",
              size=8, italic=True, wrap=True, bg=C_RED_BG, border=_border())
        r += 1

    for rec in bank_unmatched:
        idx += 1
        amt = rec.debit or rec.credit
        ws.row_dimensions[r].height = 20
        _cell(ws, r, 1, idx, align="center", size=8, bg=C_RED_BG, border=_border())
        _cell(ws, r, 2, "Unmatched", align="center", bold=True, bg=C_RED_BG, border=_border())
        _cell(ws, r, 3, "Bank", align="center", bg=C_RED_BG, border=_border())
        _cell(ws, r, 4, "—", align="center", bg=C_RED_BG, border=_border())
        _cell(ws, r, 5, str(rec.row_index), bg=C_RED_BG, border=_border())
        _cell(ws, r, 6, rec.date or rec.date_raw or "", align="center", bg=C_RED_BG, border=_border())
        _cell(ws, r, 7, "—", bg=C_RED_BG, border=_border())
        _cell(ws, r, 8, None, align="right", num_fmt=INR, bg=C_RED_BG, border=_border())
        _cell(ws, r, 9, rec.narration or "", bg=C_RED_BG, border=_border())
        _cell(ws, r, 10, amt, align="right", num_fmt=INR, bg=C_RED_BG, border=_border())
        _cell(ws, r, 11, "No Match Found", bg=C_RED_BG, border=_border())
        _cell(ws, r, 12, "—", align="center", bg=C_RED_BG, border=_border())
        _cell(ws, r, 13,
              "No corresponding ledger row was found in the Exact, Fuzzy, Memory or "
              "AI matching phases within the configured tolerances. Requires manual review.",
              size=8, italic=True, wrap=True, bg=C_RED_BG, border=_border())
        r += 1


# ── Public API ───────────────────────────────────────────────────────────────

def write_bank_recon_xlsx(
    recon_result: dict,
    gl_records:   Sequence[LedgerFormat],
    bank_records: Sequence[BankStatement],
    output_path:  Union[Path, str],
) -> None:
    """
    Write reconciliation data to a 3-sheet Excel workbook.

    Args:
        recon_result : dict returned by reconcile() in __init__.py
        gl_records   : full list of ledger records fed into reconcile()
                        (e.g. ledger_result["records"])
        bank_records : full list of bank records fed into reconcile()
                        (e.g. bank_result["records"])
        output_path  : destination .xlsx file path
    """
    path = Path(output_path)
    wb   = openpyxl.Workbook()
    wb.remove(wb.active)

    gl_by_id   = {str(r.ledger_id): r  for r in gl_records}
    bank_by_id = {str(r.row_index): r for r in bank_records}

    matches = _all_matches(recon_result)
    ledger_idx, bank_idx = _build_match_index(matches)

    _write_ledger_entries(wb, gl_records, ledger_idx)
    _write_bank_entries(wb, bank_records, bank_idx)
    _write_matched_unmatched(wb, recon_result, gl_by_id, bank_by_id)

    wb.active = wb["Matched & Unmatched"]
    wb.save(path)

    unreconciled = recon_result.get("UNRECONCILED_ITEMS", {})
    print(
        f"  Bank reconciliation saved → {path}\n"
        f"    Ledger entries: {len(gl_records)}   Bank entries: {len(bank_records)}\n"
        f"    Matched: {len(matches)}   "
        f"Unmatched (ledger/bank): "
        f"{len(unreconciled.get('ledger', []))}/{len(unreconciled.get('bank', []))}"
    )