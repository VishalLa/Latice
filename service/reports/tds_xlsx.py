from __future__ import annotations

from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from schema import TDSRegister

C_HEADER_BG  = "1E3A5F"
C_HEADER_FG  = "FFFFFF"
C_SUBHEADER  = "2E6DA4"
C_ROW_ALT    = "EBF3FB"
C_TOTAL_BG   = "FFF2CC"
C_WARN_BG    = "FBE0E0"
C_SECTION_BG = "D9E1F2"
C_BORDER     = "BFBFBF"

INR = "\u20b9#,##0.00"

def _border(style="thin", color=C_BORDER):
    s = Side(style=style, color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def _cell(ws, row, col, value=None, *, bold=False, size=9, bg=None,
          fg="000000", align="left", num_fmt=None, italic=False,
          border=None, wrap=False):
    c = ws.cell(row=row, column=col, value=value)
    c.font      = Font(name="Arial", bold=bold, size=size, color=fg, italic=italic)
    c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
    if bg:
        c.fill = PatternFill("solid", fgColor=bg)
    if num_fmt:
        c.number_format = num_fmt
    if border:
        c.border = border
    return c

def _title(ws, row: int, text: str, n_cols: int, subtitle: str = "") -> int:
    ws.row_dimensions[row].height = 32
    ws.merge_cells(f"A{row}:{get_column_letter(n_cols)}{row}")
    _cell(ws, row, 1, text, bold=True, size=14,
          bg=C_HEADER_BG, fg=C_HEADER_FG, align="center")
    row += 1
    if subtitle:
        ws.row_dimensions[row].height = 18
        ws.merge_cells(f"A{row}:{get_column_letter(n_cols)}{row}")
        _cell(ws, row, 1, subtitle, size=9,
              bg=C_SUBHEADER, fg=C_HEADER_FG, align="center", italic=True)
        row += 1
    return row + 1

def _headers(ws, row: int, hdrs: list[str]) -> int:
    ws.row_dimensions[row].height = 24
    for c, h in enumerate(hdrs, 1):
        _cell(ws, row, c, h, bold=True, size=9,
              bg=C_SUBHEADER, fg=C_HEADER_FG, align="center", border=_border())
    return row + 1

def _write_register(wb, reg: TDSRegister) -> None:
    ws = wb.create_sheet("TDS Register")
    ws.sheet_view.showGridLines = False

    col_widths = [5, 12, 12, 24, 14, 16, 14, 12, 12, 12, 14, 14, 12]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    stamp = datetime.now().strftime("%d %B %Y, %I:%M %p")
    r = _title(
        ws, 1, "TDS DEDUCTION REGISTER", len(col_widths),
        f"Period: {reg.period_start.strftime('%d-%b-%Y')} to {reg.period_end.strftime('%d-%b-%Y')}  |  "
        f"Generated: {stamp}  |  Entries: {len(reg.entries)}",
    )

    hdrs = ["#", "Date", "Section", "Deductee", "PAN", "Deductee Type",
            "Gross Amt", "TDS Base", "Rate %", "TDS Amt", "Net Payment",
            "Status", "206AA"]
    r = _headers(ws, r, hdrs)
    ws.freeze_panes = f"A{r}"

    for idx, e in enumerate(sorted(reg.entries, key=lambda x: x.date), 1):
        alt = C_ROW_ALT if idx % 2 == 0 else None
        row_vals = [
            idx, e.date.strftime("%d-%m-%Y"), e.section_code, e.deductee_name,
            e.deductee_pan or "PANNOTAVBL", e.deductee_type.value,
            e.gross_amount, e.tds_base, e.tds_rate, e.tds_amount,
            e.net_payment, e.status.value, "Yes" if e.rate_enhanced_206aa else "",
        ]
        for c, v in enumerate(row_vals, 1):
            fmt = INR if c in (7, 8, 10, 11) else None
            align = "right" if c in (7, 8, 9, 10, 11) else ("center" if c in (1, 2, 3, 9, 13) else "left")
            bg = C_WARN_BG if (c == 13 and v == "Yes") else alt
            _cell(ws, r, c, v, size=9, align=align, num_fmt=fmt, bg=bg, border=_border())
        r += 1

    _cell(ws, r, 1, "TOTAL", bold=True, size=9, bg=C_TOTAL_BG, border=_border())
    ws.merge_cells(f"A{r}:F{r}")
    _cell(ws, r, 7, reg.total_gross_amount, bold=True, size=9, bg=C_TOTAL_BG, num_fmt=INR, align="right", border=_border())
    _cell(ws, r, 10, reg.total_tds_deducted, bold=True, size=9, bg=C_TOTAL_BG, num_fmt=INR, align="right", border=_border())

def _write_section_summary(wb, reg: TDSRegister) -> None:
    ws = wb.create_sheet("Section Summary")
    ws.sheet_view.showGridLines = False
    widths = [12, 40, 14, 16, 16, 16, 16, 16]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    r = _title(ws, 1, "TDS — SECTION-WISE SUMMARY", len(widths))
    hdrs = ["Section", "Description", "Txns", "Gross Amt", "TDS Base",
            "TDS Deducted", "TDS Deposited", "Pending Deposit"]
    r = _headers(ws, r, hdrs)

    for idx, row in enumerate(reg.section_summary(), 1):
        alt = C_ROW_ALT if idx % 2 == 0 else None
        vals = [row["section_code"], row["description"], row["transaction_count"],
                row["gross_amount_total"], row["tds_base_total"], row["tds_deducted_total"],
                row["tds_deposited_total"], row["pending_deposit"]]
        for c, v in enumerate(vals, 1):
            fmt = INR if c in (4, 5, 6, 7, 8) else None
            align = "right" if c in (3, 4, 5, 6, 7, 8) else "left"
            _cell(ws, r, c, v, size=9, align=align, num_fmt=fmt, bg=alt, border=_border())
        r += 1

def _write_deductee_summary(wb, reg: TDSRegister) -> None:
    ws = wb.create_sheet("Deductee Summary")
    ws.sheet_view.showGridLines = False
    widths = [30, 16, 20, 12, 16, 16, 16]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    r = _title(ws, 1, "TDS — DEDUCTEE-WISE SUMMARY", len(widths),
               "Use this to sanity-check annual aggregate thresholds per deductee")
    hdrs = ["Deductee", "PAN", "Sections", "Txns", "Gross Amt", "TDS Deducted", "TDS Pending"]
    r = _headers(ws, r, hdrs)

    for idx, row in enumerate(reg.deductee_summary(), 1):
        alt = C_ROW_ALT if idx % 2 == 0 else None
        vals = [row["deductee_name"], row["deductee_pan"], ", ".join(row["sections"]),
                row["transaction_count"], row["gross_amount_total"],
                row["tds_deducted_total"], row["tds_pending_total"]]
        for c, v in enumerate(vals, 1):
            fmt = INR if c in (5, 6, 7) else None
            align = "right" if c in (4, 5, 6, 7) else "left"
            _cell(ws, r, c, v, size=9, align=align, num_fmt=fmt, bg=alt, border=_border())
        r += 1


def write_tds_xlsx(reg: TDSRegister, output_path: Path | str) -> None:
    path = Path(output_path)
    wb   = openpyxl.Workbook()
    wb.remove(wb.active)

    _write_register(wb, reg)
    _write_section_summary(wb, reg)
    _write_deductee_summary(wb, reg)

    wb.active = wb["TDS Register"]
    wb.save(path)
    print(
        f"  TDS register saved → {path}  "
        f"({len(reg.entries)} entries | Deducted ₹{reg.total_tds_deducted:,.2f} | "
        f"Pending ₹{reg.total_tds_pending:,.2f})"
    )
