"""
Bank Reconciliation Report Export to XLSX

Accepts the dict returned by reconcile() in __init__.py and writes a
fully-formatted Excel workbook.

Sheets produced:
  1. Summary          — one-page reconciliation statement with key counts & amounts
  2. Exact Matches    — Phase 1 results (date + amount + optional ref match)
  3. Fuzzy Matches    — Phase 2 results (13 fuzzy strategies)
  4. AI Matches       — Phase 3 semantic matches from the LLM
  5. AI Agent Queue   — items flagged for human review
  6. Unreconciled     — ledger rows and bank rows with no match found
  7. Warnings         — parse warnings collected during data extraction
"""

from __future__ import annotations

from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from schema import LedgerFormat, BankStatement

# ── Colour palette ───────────────────────────────────────────────────────────
C_HEADER_BG  = "1E3A5F"
C_HEADER_FG  = "FFFFFF"
C_SUBHEADER  = "2E6DA4"
C_ROW_ALT    = "EBF3FB"
C_TOTAL_BG   = "FFF2CC"
C_GRAND_BG   = "1E3A5F"
C_GRAND_FG   = "FFFFFF"
C_SECTION_BG = "D9E1F2"
C_BORDER     = "BFBFBF"
C_GREEN_BG   = "E2EFDA"
C_AMBER_BG   = "FFF2CC"
C_RED_BG     = "FFCCCC"
C_ORANGE_BG  = "FCE4D6"

INR = "\u20b9#,##0.00"

_CONF_COLOURS = {
    "High":   C_GREEN_BG,
    "Medium": C_AMBER_BG,
    "Low":    C_ORANGE_BG,
}


# ── Helpers ──────────────────────────────────────────────────────────────────

def _border(style="thin", color=C_BORDER):
    s = Side(style=style, color=color)
    return Border(left=s, right=s, top=s, bottom=s)


def _thick_bottom():
    t = Side(style="thin",   color=C_BORDER)
    m = Side(style="medium", color="000000")
    return Border(left=t, right=t, top=t, bottom=m)


def _cell(ws, row, col, value=None, *, bold=False, size=9, bg=None,
          fg="000000", align="left", num_fmt=None, italic=False,
          border=None, wrap=False):
    c = ws.cell(row=row, column=col, value=value)
    c.font      = Font(name="Arial", bold=bold, size=size, color=fg, italic=italic)
    c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
    if bg:
        c.fill = PatternFill("solid", fgColor=bg)
    if num_fmt:
        c.number_format = num_fmt
    if border:
        c.border = border
    return c


def _title(ws, row: int, text: str, n_cols: int, subtitle: str = "") -> int:
    ws.row_dimensions[row].height = 32
    ws.merge_cells(f"A{row}:{get_column_letter(n_cols)}{row}")
    _cell(ws, row, 1, text, bold=True, size=14,
          bg=C_HEADER_BG, fg=C_HEADER_FG, align="center")
    row += 1
    if subtitle:
        ws.row_dimensions[row].height = 18
        ws.merge_cells(f"A{row}:{get_column_letter(n_cols)}{row}")
        _cell(ws, row, 1, subtitle, size=9,
              bg=C_SUBHEADER, fg=C_HEADER_FG, align="center", italic=True)
        row += 1
    return row + 1


def _col_headers(ws, row: int, headers: list[str],
                 bg=C_SUBHEADER, fg=C_HEADER_FG) -> int:
    ws.row_dimensions[row].height = 22
    for c, h in enumerate(headers, 1):
        _cell(ws, row, c, h, bold=True, size=9,
              bg=bg, fg=fg, align="center", border=_border())
    return row + 1


def _ledger_row(rec: LedgerFormat) -> tuple:
    """Flatten a LedgerFormat into display columns."""
    amount = rec.debit_amount or rec.credit_amount
    direction = "Debit" if rec.debit_amount > 0 else "Credit"
    return (
        rec.ledger_id,
        rec.transaction_date or rec.transaction_date_raw or "",
        rec.account_name or "",
        rec.vendor_name  or "",
        rec.voucher_type or "",
        rec.reference_id or "",
        amount,
        direction,
        rec.source.value if rec.source else "",
    )


def _bank_row(rec: BankStatement) -> tuple:
    """Flatten a BankStatement into display columns."""
    amount = rec.debit or rec.credit
    direction = "Debit" if rec.debit > 0 else "Credit"
    return (
        str(rec.row_index),
        rec.date or rec.date_raw or "",
        rec.narration or "",
        rec.txn_id or "",
        amount,
        direction,
        f"₹{rec.balance:,.2f}" if rec.balance is not None else "",
    )


# ── Sheet 1 — Summary ────────────────────────────────────────────────────────

def _write_summary(wb: openpyxl.Workbook, result: dict) -> None:
    ws = wb.create_sheet("Summary")
    ws.sheet_view.showGridLines = False
    for i, w in enumerate([30, 14, 28], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    stamp = datetime.now().strftime("%d %B %Y, %I:%M %p")
    bank_name = result.get("bank_name") or "—"
    r = _title(
        ws, 1,
        "BANK RECONCILIATION STATEMENT",
        3,
        f"Bank: {bank_name}  |  Generated: {stamp}",
    )

    s = result.get("summary", {})
    rows = [
        ("RECORDS LOADED",                  None,              None),
        ("Ledger records",                  s.get("ledger_records", 0),   C_SECTION_BG),
        ("Bank statement rows",             s.get("bank_records",  0),    C_SECTION_BG),
        (None, None, None),
        ("MATCHING RESULTS",                None,              None),
        ("Phase 1 — Exact matches",         s.get("exact_matches",       0), C_GREEN_BG),
        ("Phase 2 — Fuzzy matches",         s.get("fuzzy_matches",       0), C_GREEN_BG),
        ("Phase 3 — AI semantic matches",   s.get("ai_matches",          0), C_GREEN_BG),
        (None, None, None),
        ("REVIEW QUEUE",                    None,              None),
        ("Items queued for AI agent review", s.get("ai_agent_queue",     0), C_AMBER_BG),
        (None, None, None),
        ("UNRECONCILED",                    None,              None),
        ("Unreconciled ledger rows",        s.get("unreconciled_ledger", 0), C_RED_BG),
        ("Unreconciled bank rows",          s.get("unreconciled_bank",   0), C_RED_BG),
        (None, None, None),
        ("Warnings raised",                 len(result.get("warnings", [])), C_AMBER_BG),
    ]

    for label, value, bg in rows:
        ws.row_dimensions[r].height = 20
        if label is None:
            r += 1
            continue
        if value is None:
            # Section header
            ws.merge_cells(f"A{r}:C{r}")
            _cell(ws, r, 1, label, bold=True, size=10,
                  bg=C_HEADER_BG, fg=C_HEADER_FG, border=_border())
        else:
            _cell(ws, r, 1, f"  {label}", size=10, bg=bg, border=_border())
            _cell(ws, r, 2, value, align="center", bold=True, size=11,
                  bg=bg, border=_border())
            _cell(ws, r, 3, "", bg=bg, border=_border())
        r += 1


# ── Sheet 2 — Exact Matches ──────────────────────────────────────────────────

def _write_exact_matches(wb: openpyxl.Workbook, matches: list[dict]) -> None:
    ws = wb.create_sheet("Exact Matches")
    ws.sheet_view.showGridLines = False
    for i, w in enumerate([4, 12, 12, 14, 12, 16, 12, 16], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    r = _title(ws, 1, "PHASE 1 — EXACT MATCHES", 8,
               f"{len(matches)} exact matches (date + amount, optional reference)")
    r = _col_headers(ws, r, [
        "#", "Ledger ID", "Bank Row", "Date",
        "Amount (₹)", "Reference Matched", "Ref #", "Notes",
    ])
    ws.freeze_panes = f"A{r}"

    for i, m in enumerate(matches, 1):
        alt = C_ROW_ALT if i % 2 == 0 else None
        bg  = C_GREEN_BG if m.get("reference_matched") else alt
        ws.row_dimensions[r].height = 18
        _cell(ws, r, 1, i, align="center", size=8, bg=bg, border=_border())
        _cell(ws, r, 2, str(m.get("ledger_id", "")), bg=bg, border=_border())
        _cell(ws, r, 3, str(m.get("bank_id",   "")), bg=bg, border=_border())
        _cell(ws, r, 4, str(m.get("date",       "")), align="center", bg=bg, border=_border())
        _cell(ws, r, 5, m.get("amount"), align="right", num_fmt=INR, bg=bg, border=_border())
        _cell(ws, r, 6, "Yes ✓" if m.get("reference_matched") else "No",
              align="center", bg=bg, border=_border())
        _cell(ws, r, 7, str(m.get("reference", "")), align="center", bg=bg, border=_border())
        _cell(ws, r, 8, m.get("notes", ""), size=8, italic=True, bg=bg, border=_border())
        r += 1


# ── Sheet 3 — Fuzzy Matches ──────────────────────────────────────────────────

def _write_fuzzy_matches(wb: openpyxl.Workbook, matches: list[dict]) -> None:
    ws = wb.create_sheet("Fuzzy Matches")
    ws.sheet_view.showGridLines = False
    for i, w in enumerate([4, 14, 14, 24, 12, 50], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    r = _title(ws, 1, "PHASE 2 — FUZZY MATCHES", 6,
               f"{len(matches)} matches across 13 reconciliation strategies")
    r = _col_headers(ws, r, [
        "#", "Ledger ID", "Bank Row", "Adjustment Type",
        "Confidence", "Details",
    ])
    ws.freeze_panes = f"A{r}"

    for i, m in enumerate(matches, 1):
        conf = m.get("confidence_score", "")
        bg   = _CONF_COLOURS.get(conf, C_ROW_ALT if i % 2 == 0 else None)
        ws.row_dimensions[r].height = 18
        _cell(ws, r, 1, i, align="center", size=8, bg=bg, border=_border())
        _cell(ws, r, 2, str(m.get("ledger_id",       "") or "—"), bg=bg, border=_border())
        _cell(ws, r, 3, str(m.get("bank_id",          "") or "—"), bg=bg, border=_border())
        _cell(ws, r, 4, m.get("adjustment_type", ""),
              bg=bg, border=_border())
        _cell(ws, r, 5, conf, align="center", bold=True, bg=bg, border=_border())
        _cell(ws, r, 6, m.get("details", ""), size=8, italic=True,
              bg=bg, wrap=True, border=_border())
        r += 1


# ── Sheet 4 — AI Matches ─────────────────────────────────────────────────────

def _write_ai_matches(wb: openpyxl.Workbook, matches: list[dict]) -> None:
    ws = wb.create_sheet("AI Matches")
    ws.sheet_view.showGridLines = False
    for i, w in enumerate([4, 12, 12, 12, 14, 14, 14, 12, 40], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    r = _title(ws, 1, "PHASE 3 — AI SEMANTIC MATCHES", 9,
               f"{len(matches)} semantic matches identified by the LLM")
    r = _col_headers(ws, r, [
        "#", "Ledger ID", "Bank Row",
        "Ledger Date", "Ledger Account", "Ledger Amt (₹)",
        "Bank Narration", "Confidence", "Reason",
    ])
    ws.freeze_panes = f"A{r}"

    for i, m in enumerate(matches, 1):
        conf = m.get("confidence", "")
        bg   = _CONF_COLOURS.get(conf, C_ROW_ALT if i % 2 == 0 else None)

        l_item: LedgerFormat  | None = m.get("ledger_item")
        b_item: BankStatement | None = m.get("bank_item")

        l_date   = l_item.transaction_date if l_item else ""
        l_name   = l_item.account_name     if l_item else ""
        l_amt    = (l_item.debit_amount or l_item.credit_amount) if l_item else None
        b_narr   = b_item.narration        if b_item else ""

        ws.row_dimensions[r].height = 18
        _cell(ws, r, 1, i, align="center", size=8, bg=bg, border=_border())
        _cell(ws, r, 2, m.get("ledger_id",      ""), bg=bg, border=_border())
        _cell(ws, r, 3, m.get("bank_row_index", ""), bg=bg, border=_border())
        _cell(ws, r, 4, str(l_date),   align="center", bg=bg, border=_border())
        _cell(ws, r, 5, l_name,        bg=bg, border=_border())
        _cell(ws, r, 6, l_amt,         align="right", num_fmt=INR, bg=bg, border=_border())
        _cell(ws, r, 7, b_narr,        bg=bg, border=_border())
        _cell(ws, r, 8, conf,          align="center", bold=True, bg=bg, border=_border())
        _cell(ws, r, 9, m.get("reason", ""), size=8, italic=True,
              wrap=True, bg=bg, border=_border())
        r += 1


# ── Sheet 5 — AI Agent Queue ─────────────────────────────────────────────────

def _write_ai_agent_queue(wb: openpyxl.Workbook, queue: list[dict]) -> None:
    ws = wb.create_sheet("AI Agent Queue")
    ws.sheet_view.showGridLines = False
    for i, w in enumerate([4, 14, 14, 24, 12, 50], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    r = _title(ws, 1, "AI AGENT REVIEW QUEUE", 6,
               f"{len(queue)} items require human / agent review before finalisation")
    r = _col_headers(ws, r, [
        "#", "Ledger ID", "Bank Row / Bank ID",
        "Reason for Review", "Confidence", "Details",
    ])
    ws.freeze_panes = f"A{r}"

    for i, item in enumerate(queue, 1):
        # Items can come from fuzzy low-confidence or from ai_result
        conf    = item.get("confidence", item.get("confidence_score", ""))
        adj     = item.get("adjustment_type", item.get("reason", ""))
        details = item.get("details",  item.get("reason", ""))
        lid     = str(item.get("ledger_id",      item.get("ledger_id",     "")) or "—")
        bid     = str(item.get("bank_row_index",  item.get("bank_id",       "")) or "—")

        bg = _CONF_COLOURS.get(conf, C_AMBER_BG)
        ws.row_dimensions[r].height = 20
        _cell(ws, r, 1, i, align="center", size=8, bg=bg, border=_border())
        _cell(ws, r, 2, lid, bg=bg, border=_border())
        _cell(ws, r, 3, bid, bg=bg, border=_border())
        _cell(ws, r, 4, adj, bg=bg, border=_border())
        _cell(ws, r, 5, conf, align="center", bold=True, bg=bg, border=_border())
        _cell(ws, r, 6, details, size=8, italic=True, wrap=True, bg=bg, border=_border())
        r += 1


# ── Sheet 6 — Unreconciled ───────────────────────────────────────────────────

def _write_unreconciled(wb: openpyxl.Workbook, unreconciled: dict) -> None:
    ws = wb.create_sheet("Unreconciled")
    ws.sheet_view.showGridLines = False
    for i, w in enumerate([4, 12, 12, 28, 22, 14, 16, 12, 16, 12], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ledger_items: list[LedgerFormat]  = unreconciled.get("ledger", [])
    bank_items:   list[BankStatement] = unreconciled.get("bank",   [])
    total = len(ledger_items) + len(bank_items)

    r = _title(ws, 1, "UNRECONCILED ITEMS", 10,
               f"{total} items remain after all three matching phases")

    # ── Ledger section ────────────────────────────────────────────────────────
    if ledger_items:
        ws.row_dimensions[r].height = 20
        ws.merge_cells(f"A{r}:J{r}")
        _cell(ws, r, 1, f"  LEDGER UNRECONCILED ({len(ledger_items)} items)",
              bold=True, size=10, bg=C_HEADER_BG, fg=C_HEADER_FG, border=_border())
        r += 1

        r = _col_headers(ws, r, [
            "#", "Ledger ID", "Date", "Account Name",
            "Vendor", "Voucher Type", "Debit (₹)", "Credit (₹)",
            "Reference", "Source",
        ])

        for i, rec in enumerate(ledger_items, 1):
            alt = C_ROW_ALT if i % 2 == 0 else None
            ws.row_dimensions[r].height = 18
            _cell(ws, r,  1, i, align="center", size=8, bg=alt or C_RED_BG, border=_border())
            _cell(ws, r,  2, rec.ledger_id or "",          bg=alt or C_RED_BG, border=_border())
            _cell(ws, r,  3, rec.transaction_date or "",   align="center", bg=alt or C_RED_BG, border=_border())
            _cell(ws, r,  4, rec.account_name or "",       bg=alt or C_RED_BG, border=_border())
            _cell(ws, r,  5, rec.vendor_name or "",        bg=alt or C_RED_BG, border=_border())
            _cell(ws, r,  6, rec.voucher_type or "",       size=8, bg=alt or C_RED_BG, border=_border())
            _cell(ws, r,  7, rec.debit_amount  or None,   align="right", num_fmt=INR, bg=alt or C_RED_BG, border=_border())
            _cell(ws, r,  8, rec.credit_amount or None,   align="right", num_fmt=INR, bg=alt or C_RED_BG, border=_border())
            _cell(ws, r,  9, rec.reference_id or "",      align="center", bg=alt or C_RED_BG, border=_border())
            _cell(ws, r, 10, rec.source.value if rec.source else "", size=8, bg=alt or C_RED_BG, border=_border())
            r += 1
        r += 1

    # ── Bank section ─────────────────────────────────────────────────────────
    if bank_items:
        ws.row_dimensions[r].height = 20
        ws.merge_cells(f"A{r}:J{r}")
        _cell(ws, r, 1, f"  BANK UNRECONCILED ({len(bank_items)} items)",
              bold=True, size=10, bg=C_HEADER_BG, fg=C_HEADER_FG, border=_border())
        r += 1

        r = _col_headers(ws, r, [
            "#", "Row Index", "Date", "Narration",
            "Txn ID", "Debit (₹)", "Credit (₹)", "Balance (₹)",
            "Bank", "",
        ])

        for i, rec in enumerate(bank_items, 1):
            alt = C_ROW_ALT if i % 2 == 0 else None
            ws.row_dimensions[r].height = 18
            _cell(ws, r,  1, i, align="center", size=8, bg=alt or C_RED_BG, border=_border())
            _cell(ws, r,  2, str(rec.row_index), align="center", bg=alt or C_RED_BG, border=_border())
            _cell(ws, r,  3, rec.date or "",     align="center", bg=alt or C_RED_BG, border=_border())
            _cell(ws, r,  4, rec.narration or "", bg=alt or C_RED_BG, border=_border())
            _cell(ws, r,  5, rec.txn_id or "",   align="center", bg=alt or C_RED_BG, border=_border())
            _cell(ws, r,  6, rec.debit  or None, align="right", num_fmt=INR, bg=alt or C_RED_BG, border=_border())
            _cell(ws, r,  7, rec.credit or None, align="right", num_fmt=INR, bg=alt or C_RED_BG, border=_border())
            _cell(ws, r,  8, rec.balance,         align="right", num_fmt=INR, bg=alt or C_RED_BG, border=_border())
            _cell(ws, r,  9, rec.bank_name or "", size=8, bg=alt or C_RED_BG, border=_border())
            _cell(ws, r, 10, "",                  bg=alt or C_RED_BG, border=_border())
            r += 1


# ── Sheet 7 — Warnings ───────────────────────────────────────────────────────

def _write_warnings(wb: openpyxl.Workbook, warnings: list) -> None:
    if not warnings:
        return
    ws = wb.create_sheet("Warnings ⚠")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 90

    r = _title(ws, 1, "WARNINGS & PARSE NOTES", 2,
               f"{len(warnings)} warning(s) raised during processing")
    r = _col_headers(ws, r, ["#", "Warning Message"])

    for i, w in enumerate(warnings, 1):
        alt = C_ROW_ALT if i % 2 == 0 else None
        ws.row_dimensions[r].height = 18
        _cell(ws, r, 1, i, align="center", size=8,
              bg=alt or C_AMBER_BG, border=_border())
        _cell(ws, r, 2, str(w), size=9, wrap=True,
              bg=alt or C_AMBER_BG, border=_border())
        r += 1


# ── Public API ───────────────────────────────────────────────────────────────

def write_bank_recon_xlsx(
    recon_result: dict,
    output_path:  Path | str,
) -> None:
    """
    Write all reconciliation data to an Excel workbook.

    Args:
        recon_result : dict returned by reconcile() in __init__.py
        output_path  : destination .xlsx file path
    """
    path = Path(output_path)
    wb   = openpyxl.Workbook()
    wb.remove(wb.active)

    exact_matches = recon_result.get("EXACT_MATCHES",    [])
    fuzzy_matches = recon_result.get("FUZZY_MATCHES",    [])
    ai_matches    = recon_result.get("AI_MATCHES",       [])
    ai_queue      = recon_result.get("AI_AGENT",         [])
    unreconciled  = recon_result.get("UNRECONCILED_ITEMS", {})
    warnings      = recon_result.get("warnings",         [])

    _write_summary(wb, recon_result)
    _write_exact_matches(wb, exact_matches)
    _write_fuzzy_matches(wb, fuzzy_matches)
    _write_ai_matches(wb, ai_matches)
    _write_ai_agent_queue(wb, ai_queue)
    _write_unreconciled(wb, unreconciled)
    _write_warnings(wb, warnings)

    wb.active = wb["Summary"]
    wb.save(path)

    s = recon_result.get("summary", {})
    print(
        f"  Bank reconciliation saved → {path}\n"
        f"    Exact: {s.get('exact_matches',0)}  "
        f"Fuzzy: {s.get('fuzzy_matches',0)}  "
        f"AI: {s.get('ai_matches',0)}  "
        f"Queue: {s.get('ai_agent_queue',0)}  "
        f"Unreconciled L/B: "
        f"{s.get('unreconciled_ledger',0)}/{s.get('unreconciled_bank',0)}"
    )
    