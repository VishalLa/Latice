"""
ledger_xlsx.py — General Ledger Export to XLSX

Writes all data that lives inside a GeneralLedger object into a formatted
Excel workbook.  Call write_ledger_xlsx() after build_ledger() returns.

Sheets produced:
  1. Trial Balance      — closing debit / credit per account, grouped by AccountGroup
  2. Ledger Accounts    — T-account style posting history for every account
  3. Cash Book          — combined Cash + Bank running balance
"""

from __future__ import annotations

from collections import defaultdict
from datetime import datetime
from pathlib import Path
from typing import Optional

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from schema import AccountGroup, CashBookLine
from .ledger_xlsx import GeneralLedger, TrialBalance, CashBookLine, trial_balance, extract_cash_book

# ── Colour palette ───────────────────────────────────────────────────────────
C_HEADER_BG  = "1E3A5F"
C_HEADER_FG  = "FFFFFF"
C_SUBHEADER  = "2E6DA4"
C_ROW_ALT    = "EBF3FB"
C_TOTAL_BG   = "FFF2CC"
C_GRAND_BG   = "1E3A5F"
C_GRAND_FG   = "FFFFFF"
C_SECTION_BG = "D9E1F2"
C_BORDER     = "BFBFBF"
C_DEBIT_BG   = "FFF0E0"
C_CREDIT_BG  = "E0F0FF"

INR = "\u20b9#,##0.00"

# Schedule III display order (liabilities → assets → income → expenses)
_GROUP_ORDER = [
    AccountGroup.CAPITAL_ACCOUNT.value,
    AccountGroup.RESERVES_SURPLUS.value,
    AccountGroup.LOANS_LIABILITY.value,
    AccountGroup.CURRENT_LIABILITIES.value,
    AccountGroup.SUNDRY_CREDITORS.value,
    AccountGroup.DUTIES_TAXES.value,
    AccountGroup.FIXED_ASSETS.value,
    AccountGroup.INVESTMENTS.value,
    AccountGroup.CURRENT_ASSETS.value,
    AccountGroup.SUNDRY_DEBTORS.value,
    AccountGroup.CASH_IN_HAND.value,
    AccountGroup.BANK_ACCOUNTS.value,
    AccountGroup.STOCK_IN_HAND.value,
    AccountGroup.LOANS_ADVANCES.value,
    AccountGroup.SALES_ACCOUNTS.value,
    AccountGroup.DIRECT_INCOME.value,
    AccountGroup.INDIRECT_INCOME.value,
    AccountGroup.PURCHASE_ACCOUNTS.value,
    AccountGroup.DIRECT_EXPENSES.value,
    AccountGroup.INDIRECT_EXPENSES.value,
]


# ── Helpers ──────────────────────────────────────────────────────────────────

def _border(style="thin", color=C_BORDER):
    s = Side(style=style, color=color)
    return Border(left=s, right=s, top=s, bottom=s)


def _thick_bottom():
    t = Side(style="thin",   color=C_BORDER)
    m = Side(style="medium", color="000000")
    return Border(left=t, right=t, top=t, bottom=m)


def _cell(ws, row, col, value=None, *, bold=False, size=9, bg=None,
          fg="000000", align="left", num_fmt=None, italic=False,
          border=None, wrap=False):
    c = ws.cell(row=row, column=col, value=value)
    c.font      = Font(name="Arial", bold=bold, size=size, color=fg, italic=italic)
    c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
    if bg:
        c.fill = PatternFill("solid", fgColor=bg)
    if num_fmt:
        c.number_format = num_fmt
    if border:
        c.border = border
    return c


def _title(ws, row: int, text: str, n_cols: int, subtitle: str = "") -> int:
    ws.row_dimensions[row].height = 32
    ws.merge_cells(f"A{row}:{get_column_letter(n_cols)}{row}")
    _cell(ws, row, 1, text, bold=True, size=14,
          bg=C_HEADER_BG, fg=C_HEADER_FG, align="center")
    row += 1
    if subtitle:
        ws.row_dimensions[row].height = 18
        ws.merge_cells(f"A{row}:{get_column_letter(n_cols)}{row}")
        _cell(ws, row, 1, subtitle, size=9,
              bg=C_SUBHEADER, fg=C_HEADER_FG, align="center", italic=True)
        row += 1
    return row + 1


def _section_header(ws, row: int, text: str, n_cols: int) -> int:
    ws.row_dimensions[row].height = 20
    ws.merge_cells(f"A{row}:{get_column_letter(n_cols)}{row}")
    _cell(ws, row, 1, text, bold=True, size=9, bg=C_SECTION_BG)
    return row + 1


# ── Sheet 1 — Trial Balance ──────────────────────────────────────────────────

def _write_trial_balance(wb: openpyxl.Workbook, tb: TrialBalance) -> None:
    ws = wb.create_sheet("Trial Balance")
    ws.sheet_view.showGridLines = False

    for i, w in enumerate([4, 35, 22, 16, 16], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    bal_label = (
        "BALANCED ✓"
        if tb.is_balanced
        else f"DIFFERENCE ₹{tb.difference:,.2f} — CHECK ENTRIES"
    )
    r = _title(
        ws, 1,
        "TRIAL BALANCE",
        5,
        f"As on {tb.as_on.strftime('%d %B %Y')}  |  {bal_label}",
    )

    hdrs = ["#", "Account Name", "Account Group", "Debit (₹)", "Credit (₹)"]
    ws.row_dimensions[r].height = 24
    for c, h in enumerate(hdrs, 1):
        _cell(ws, r, c, h, bold=True, size=9,
              bg=C_SUBHEADER, fg=C_HEADER_FG, align="center", border=_border())
    r += 1
    ws.freeze_panes = f"A{r}"

    grouped: dict[str, list] = defaultdict(list)
    for line in tb.lines:
        grouped[line.group].append(line)

    all_groups = _GROUP_ORDER + [g for g in grouped if g not in _GROUP_ORDER]
    idx = 1
    for grp in all_groups:
        lines = grouped.get(grp)
        if not lines:
            continue
        r = _section_header(ws, r, grp, 5)
        for line in lines:
            alt = C_ROW_ALT if idx % 2 == 0 else None
            ws.row_dimensions[r].height = 18
            _cell(ws, r, 1, idx, align="center", size=8, bg=alt, border=_border())
            _cell(ws, r, 2, line.account, bg=alt, border=_border())
            _cell(ws, r, 3, line.group, size=8, bg=alt, border=_border())
            _cell(ws, r, 4, line.closing_debit  or None, align="right",
                  num_fmt=INR, bg=alt, border=_border())
            _cell(ws, r, 5, line.closing_credit or None, align="right",
                  num_fmt=INR, bg=alt, border=_border())
            r += 1
            idx += 1

    ws.row_dimensions[r].height = 24
    ws.merge_cells(f"A{r}:C{r}")
    _cell(ws, r, 1, "TOTAL", bold=True, size=11,
          bg=C_GRAND_BG, fg=C_GRAND_FG, align="right", border=_thick_bottom())
    _cell(ws, r, 4, tb.total_debit,  bold=True, size=11,
          bg=C_GRAND_BG, fg=C_GRAND_FG, align="right",
          num_fmt=INR, border=_thick_bottom())
    _cell(ws, r, 5, tb.total_credit, bold=True, size=11,
          bg=C_GRAND_BG, fg=C_GRAND_FG, align="right",
          num_fmt=INR, border=_thick_bottom())


# ── Sheet 2 — Ledger Accounts ────────────────────────────────────────────────

def _write_ledger_accounts(wb: openpyxl.Workbook, gl: GeneralLedger) -> None:
    ws = wb.create_sheet("Ledger Accounts")
    ws.sheet_view.showGridLines = False

    for i, w in enumerate([4, 14, 35, 20, 14, 14, 18], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    stamp = datetime.now().strftime("%d %B %Y, %I:%M %p")
    r = _title(
        ws, 1,
        "GENERAL LEDGER — ALL ACCOUNTS",
        7,
        f"Generated: {stamp}  |  Account-wise postings with running balance",
    )

    for ledger_acc in gl.accounts:
        if not ledger_acc.postings:
            continue

        bal_amt, bal_side = ledger_acc.closing_balance

        # Account banner
        ws.row_dimensions[r].height = 22
        ws.merge_cells(f"A{r}:G{r}")
        _cell(ws, r, 1,
              f"  {ledger_acc.name}  [{ledger_acc.group.value}]"
              f"  —  Closing Balance: ₹{bal_amt:,.2f} {bal_side}",
              bold=True, size=10,
              bg=C_HEADER_BG, fg=C_HEADER_FG, border=_border())
        r += 1

        # Column headers
        hdrs = ["#", "Date", "Particulars", "Voucher Type", "Dr (₹)", "Cr (₹)", "Balance"]
        ws.row_dimensions[r].height = 18
        for c, h in enumerate(hdrs, 1):
            _cell(ws, r, c, h, bold=True, size=8,
                  bg=C_SUBHEADER, fg=C_HEADER_FG, align="center", border=_border())
        r += 1

        for i, p in enumerate(ledger_acc.postings, 1):
            alt = C_ROW_ALT if i % 2 == 0 else None
            ws.row_dimensions[r].height = 18
            _cell(ws, r, 1, i, size=8, align="center", bg=alt, border=_border())
            _cell(ws, r, 2, p.date.strftime("%d-%m-%Y"), bg=alt, border=_border())
            _cell(ws, r, 3, p.particulars, bg=alt, border=_border())
            _cell(ws, r, 4, p.voucher_type, size=8, bg=alt, border=_border())
            _cell(ws, r, 5, p.dr_amount or None, align="right",
                  num_fmt=INR, bg=alt, border=_border())
            _cell(ws, r, 6, p.cr_amount or None, align="right",
                  num_fmt=INR, bg=alt, border=_border())
            _cell(ws, r, 7, f"₹{p.balance:,.2f} {p.balance_side}",
                  align="right", bg=alt, border=_border())
            r += 1

        # Closing balance row
        ws.row_dimensions[r].height = 20
        ws.merge_cells(f"A{r}:D{r}")
        _cell(ws, r, 1, "Closing Balance", bold=True, size=9,
              bg=C_TOTAL_BG, align="right", border=_thick_bottom())
        _cell(ws, r, 5, ledger_acc.total_debits,  bold=True, size=9,
              bg=C_TOTAL_BG, align="right", num_fmt=INR, border=_thick_bottom())
        _cell(ws, r, 6, ledger_acc.total_credits, bold=True, size=9,
              bg=C_TOTAL_BG, align="right", num_fmt=INR, border=_thick_bottom())
        _cell(ws, r, 7, f"₹{bal_amt:,.2f} {bal_side}", bold=True, size=9,
              bg=C_TOTAL_BG, align="right", border=_thick_bottom())
        r += 2


# ── Sheet 3 — Cash Book ──────────────────────────────────────────────────────

def _write_cash_book(wb: openpyxl.Workbook, lines: list[CashBookLine]) -> None:
    ws = wb.create_sheet("Cash Book")
    ws.sheet_view.showGridLines = False

    for i, w in enumerate([4, 12, 35, 18, 12, 14, 14, 14], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    r = _title(
        ws, 1,
        "CASH BOOK",
        8,
        f"Combined Cash & Bank — {len(lines)} transactions",
    )

    hdrs = ["#", "Date", "Particulars", "Voucher Type",
            "Account", "Receipts (₹)", "Payments (₹)", "Balance (₹)"]
    ws.row_dimensions[r].height = 24
    for c, h in enumerate(hdrs, 1):
        _cell(ws, r, c, h, bold=True, size=9,
              bg=C_SUBHEADER, fg=C_HEADER_FG, align="center", border=_border())
    r += 1
    ws.freeze_panes = f"A{r}"

    tot_receipts = tot_payments = 0.0
    for i, line in enumerate(lines, 1):
        alt = C_ROW_ALT if i % 2 == 0 else None
        tot_receipts += line.receipts
        tot_payments += line.payments
        ws.row_dimensions[r].height = 18
        _cell(ws, r, 1, i, align="center", size=8, bg=alt, border=_border())
        _cell(ws, r, 2, line.date.strftime("%d-%m-%Y"), bg=alt, border=_border())
        _cell(ws, r, 3, line.particulars, bg=alt, border=_border())
        _cell(ws, r, 4, line.voucher_type, size=8, bg=alt, border=_border())
        _cell(ws, r, 5, line.account_type, align="center", bg=alt, border=_border())
        _cell(ws, r, 6, line.receipts or None, align="right",
              num_fmt=INR, bg=alt or C_CREDIT_BG, border=_border())
        _cell(ws, r, 7, line.payments or None, align="right",
              num_fmt=INR, bg=alt or C_DEBIT_BG, border=_border())
        _cell(ws, r, 8, line.balance, align="right",
              num_fmt=INR, bg=alt, border=_border())
        r += 1

    ws.row_dimensions[r].height = 24
    ws.merge_cells(f"A{r}:E{r}")
    _cell(ws, r, 1, "TOTAL", bold=True, size=11,
          bg=C_GRAND_BG, fg=C_GRAND_FG, align="right", border=_thick_bottom())
    _cell(ws, r, 6, tot_receipts, bold=True, size=11,
          bg=C_GRAND_BG, fg=C_GRAND_FG, align="right",
          num_fmt=INR, border=_thick_bottom())
    _cell(ws, r, 7, tot_payments, bold=True, size=11,
          bg=C_GRAND_BG, fg=C_GRAND_FG, align="right",
          num_fmt=INR, border=_thick_bottom())
    _cell(ws, r, 8, "", bg=C_GRAND_BG, border=_thick_bottom())


# ── Public API ───────────────────────────────────────────────────────────────

def write_ledger_xlsx(
    gl:          GeneralLedger,
    output_path: Path | str,
    as_on=None,
) -> None:
    """
    Write Trial Balance, Ledger Accounts and Cash Book to an Excel workbook.

    Args:
        gl          : GeneralLedger populated by build_ledger()
        output_path : destination .xlsx path
        as_on       : date for the trial balance (defaults to today)
    """
    path = Path(output_path)
    wb   = openpyxl.Workbook()
    wb.remove(wb.active)

    tb   = trial_balance(gl, as_on=as_on)
    cb   = extract_cash_book(gl)

    _write_trial_balance(wb, tb)
    _write_ledger_accounts(wb, gl)
    _write_cash_book(wb, cb)

    wb.active = wb["Trial Balance"]
    wb.save(path)

    n_accs = len([a for a in gl.accounts if a.postings])
    bal_label = "BALANCED ✓" if tb.is_balanced else f"DIFF ₹{tb.difference:,.2f}"
    print(
        f"  Ledger report saved → {path}  "
        f"({n_accs} accounts | TB {bal_label} | {len(cb)} cash-book lines)"
    )